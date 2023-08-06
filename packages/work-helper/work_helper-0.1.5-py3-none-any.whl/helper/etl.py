import os
import argparse

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.read_only import ReadOnlyCell

from helper.base.writer import EnhancedWriter
from helper.base.logger import log
from helper.plugin.etl_base import Args, ModelFile, PluginBase
from helper.plugin.etl_postgre import Plugin as PostgrePlugin

__version__ = "0.0.9"


def find_model_file():
    cwd = os.getcwd()
    entrys = os.listdir(cwd)

    for entry in entrys:
        file = os.path.join(cwd, entry)
        if (
            os.path.isfile(file)
            and file.endswith(".xlsx")
            and not entry.startswith("~$")
        ):
            return file
    return None


def is_model_sheet(sheet: Worksheet):
    if sheet.max_row < 13:
        return False
    A13: ReadOnlyCell = sheet["A13"]
    if A13.value != "目标字段英文名":
        return False
    return True


def load_model_file(args: Args):

    book: Workbook = load_workbook(
        args.file, read_only=True, data_only=True, keep_vba=False
    )

    for sheet_name in book.sheetnames:
        sheet: Worksheet = book[sheet_name]
        if is_model_sheet(sheet):
            model = ModelFile(sheet, args)
            yield model


def generate_etl_file(model: ModelFile, plugin: PluginBase):
    file = os.path.join(model.args.dest, f"{model.target_table_name}.sql")
    with EnhancedWriter(file) as writer:
        plugin.write_meta_comments(model, writer)
        writer.writeln()
        plugin.write_etl_script(model, writer)
    log.info(f"生成ETL脚本：{model.sheet_name} > {file}")


def generate_ddl_file(model: ModelFile, plugin: PluginBase):
    file = os.path.join(model.args.dest, f"{model.target_table_name}.ddl.sql")
    with EnhancedWriter(file) as writer:
        plugin.write_ddl_script(model, writer)
    log.info(f"生成DDL脚本：{model.sheet_name} > {file}")


def load_plugin(args: Args):
    match args.target:
        case "postgre":
            return PostgrePlugin()
        case _:
            return None


def main():

    parser = argparse.ArgumentParser(description="ETL脚本生成程序")
    parser.add_argument(
        "--version", action="version", version=__version__, help="显示程序版本号"
    )
    parser.add_argument(
        "file",
        type=str,
        nargs="?",
        default=None,
        help="Excel模型文件(*.xlsx)，如果不指定会从当前工作目录自动寻找第一个模型文件进行处理。",
    )
    parser.add_argument("--dest", default=".", type=str, help="设定文件生成目录，默认为当前工作目录。")
    parser.add_argument(
        "--target", default="postgre", type=str, help="设定生成目标，默认postgre。"
    )

    args: Args = parser.parse_args()

    log.info(f"{parser.description} {__version__}")

    if args.file is None:
        args.file = find_model_file()

    if args.file is None:
        log.error("未找到Excel模型文件")
        exit(1)

    os.makedirs(args.dest, exist_ok=True)

    for model in load_model_file(args):
        plugin = load_plugin(args)
        generate_etl_file(model, plugin)
        generate_ddl_file(model, plugin)


if __name__ == "__main__":
    main()
