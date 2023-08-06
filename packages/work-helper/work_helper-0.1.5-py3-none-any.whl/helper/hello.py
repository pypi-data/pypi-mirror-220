import os
import sys

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from sqlalchemy import create_engine, MetaData, Table
from helper.base.logger import log

# add hidden import for pyinstaller.
from cryptography.hazmat.primitives.kdf import pbkdf2


FILES = [
    "cbondanalysiscsi.xlsx",
    "cbondanalysisshc.xlsx",
    "cbondanalysiscnbd.xlsx",
    "windcustomcode.xlsx",
]

DIR = "C:\\work\\temp"

CONNECTION_URL = (
    "oracle+oracledb://odswind:fds#2020@10.10.24.250:1521?service_name=fdsuat"
)


class DBSink:
    def __init__(self) -> None:
        self.engine = create_engine(CONNECTION_URL)
        self.meta_data = MetaData()

    def get_table(self, table_name: str) -> Table:
        table = self.meta_data.tables.get(table_name)
        if table is not None:
            return table
        return Table(table_name, self.meta_data, autoload_with=self.engine)

    def make_delete_expression(self, table: Table, row: dict):
        exp = table.delete()
        keys = table.primary_key.columns.keys()
        if len(keys) == 0:
            return None
        for key in keys:
            exp = exp.where(table.c[key] == row[key])
        return exp

    def save(self, table_name, rows):
        table = self.get_table(table_name=table_name)
        with self.engine.begin() as con:
            for row in rows:
                con.execute(self.make_delete_expression(table, row))
                con.execute(table.insert(), row)


def job(file, table: str):
    book: Workbook = load_workbook(file, read_only=True, data_only=True)
    sink = DBSink()

    columns: list[str] = None
    rows = []
    total_count = 0

    for sheet_name in book.sheetnames:
        sheet: Worksheet = book[sheet_name]
        for row in sheet.iter_rows():
            if columns is None:
                columns = list(map(lambda x: x.value, row))
                continue
            obj = {}
            for i in range(len(columns)):
                obj[columns[i]] = row[i].value
            rows.append(obj)

            count = len(rows)
            if count >= 3000:
                sink.save(table, rows)
                rows.clear()
                total_count = total_count + count
                print(total_count)

    if len(rows) > 0:
        sink.save(table, rows)
        total_count = total_count + len(rows)
        print(total_count)


def main():
    file_name = sys.argv[1]
    table_name = sys.argv[2]
    log.info(f"start import data from {file_name} to oracle table {table_name}")
    job(os.path.join(DIR, file_name), table_name)
    log.info("import job completed!")


if __name__ == "__main__":
    main()
