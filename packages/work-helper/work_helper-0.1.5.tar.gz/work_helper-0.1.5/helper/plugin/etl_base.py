from typing import Protocol
from helper.base.writer import EnhancedWriter
from openpyxl.worksheet.worksheet import Worksheet
from abc import ABC, abstractmethod


class Args(Protocol):
    file: str
    dest: str
    target: str


class DataRow:
    def __init__(self, data) -> None:
        self.data = data

    def __getitem__(self, key):
        return self.data[key]

    def is_main_table(self) -> bool:
        return self.data[6] == "MAIN TABLE"

    def is_join(self) -> bool:
        return self.data[6] is not None and not self.is_main_table()

    @property
    def join_type(self) -> str:
        return self.data[6]

    @property
    def source_table(self) -> str:
        """源表名称"""
        return self.data[10]

    @property
    def column_name(self) -> str:
        """目标表列名"""
        return self.data[0]

    @property
    def column_cn_name(self) -> str:
        """目标表中文列名"""
        return self.data[1]

    @property
    def mapping_rule(self) -> str:
        """映射规则"""
        return self.data[15]

    @property
    def source_table_id(self) -> str:
        """源表别名"""
        return self.data[9]

    @property
    def join_on_condition(self) -> str:
        """JOIN表的ON条件"""
        return self.data[7]

    @property
    def where_condition(self) -> str:
        """主表行的SQL查询条件"""
        return self.data[19]

    @property
    def column_is_pk(self) -> bool:
        """是否主键"""
        return self.data[4] == "Y"

    @property
    def column_type(self) -> str:
        """列类型"""
        return self.data[2]

    @property
    def column_null_str(self) -> str:
        """是否非空字符串"""
        if self.column_is_pk:
            return "NOT NULL"
        return "NULL"


class DataGroup:
    def __init__(self, group_id) -> None:
        self.data: list[DataRow] = []
        self.group_id = group_id

    def main_table(self):
        return next(x for x in self.data if x.is_main_table())


class ModelFile:
    def __init__(self, sheet: Worksheet, args: Args) -> None:
        self.sheet_name: str = sheet.title
        self.args = args
        self.meta: list[tuple[str, str]] = []
        for row in range(2, 12):
            self.meta.append((sheet.cell(row, 1).value, sheet.cell(row, 2).value))

        self.target_table_name = self.meta[2][1]
        self.target_table_cn_name = self.meta[1][1]

        info = self._table_info()
        self.table_name = info[1]
        self.schema = info[0]

        groups = {}

        for v in sheet.iter_rows(min_row=14, values_only=True):
            if v[0] is not None:
                gid = v[5]
                if gid not in groups:
                    groups[gid] = DataGroup(gid)
                group: DataGroup = groups[gid]
                group.data.append(DataRow(v))

        self.data: list[DataGroup] = list(groups.values())

    def _table_info(self) -> tuple[str | None, str]:
        array = self.target_table_name.split(".")
        if len(array) == 1:
            return None, self.target_table_name
        return array[0], array[1]

    def get_primary_keys(self):
        group = self.data[0]
        keys = filter(lambda x: x.column_is_pk, group.data)
        return list(map(lambda x: x.column_name, keys))


class PluginBase(ABC):
    def write_meta_comments(self, model: ModelFile, writer: EnhancedWriter):
        for kv in model.meta:
            key = str(kv[0]).replace("\n", " ")
            vs = str(kv[1]).split("\n")
            writer.write(f"-- {key.ljust(10,chr(12288))}: {vs[0]}\n")
            if len(vs) > 1:
                for v in vs[1:]:
                    writer.write(f"-- {''.ljust(10,chr(12288))}  {v}\n")

    @abstractmethod
    def write_etl_script(self, model: ModelFile, writer: EnhancedWriter):
        pass

    @abstractmethod
    def write_ddl_script(self, model: ModelFile, writer: EnhancedWriter):
        pass
