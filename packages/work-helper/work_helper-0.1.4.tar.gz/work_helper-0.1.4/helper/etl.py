import argparse
from helper import EnhancedWriter
from helper.base.logger import log
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.read_only import ReadOnlyCell
import os
from typing import Protocol
from io import StringIO

__version__ = "0.0.8"


class Args(Protocol):
    file: str
    dest: str


class DataRow:
    def __init__(self, data) -> None:
        self.data = data

    def __getitem__(self, key):
        return self.data[key]

    def is_main_table(self) -> bool:
        return self.data[6] == "MAIN TABLE"

    def is_join(self) -> bool:
        return self.data[6] is not None and not self.is_main_table()

    @property
    def join_type(self) -> str:
        return self.data[6]

    @property
    def source_table(self) -> str:
        """源表名称"""
        return self.data[10]

    @property
    def column_name(self) -> str:
        """目标表列名"""
        return self.data[0]

    @property
    def column_cn_name(self) -> str:
        """目标表中文列名"""
        return self.data[1]

    @property
    def mapping_rule(self) -> str:
        """映射规则"""
        return self.data[15]

    @property
    def source_table_id(self) -> str:
        """源表别名"""
        return self.data[9]

    @property
    def join_on_condition(self) -> str:
        """JOIN表的ON条件"""
        return self.data[7]

    @property
    def where_condition(self) -> str:
        """主表行的SQL查询条件"""
        return self.data[19]

    @property
    def column_is_pk(self) -> bool:
        """是否主键"""
        return self.data[4] == "Y"

    @property
    def column_type(self) -> str:
        """列类型"""
        return self.data[2]

    @property
    def column_null_str(self) -> str:
        """是否非空字符串"""
        if self.column_is_pk:
            return "NOT NULL"
        return "NULL"


class DataGroup:
    def __init__(self, group_id) -> None:
        self.data: list[DataRow] = []
        self.group_id = group_id

    def main_table(self):
        return next(x for x in self.data if x.is_main_table())

    def write_sql(self, w: EnhancedWriter):
        w.write("SELECT\n")
        for i in range(len(self.data)):
            d = self.data[i]
            w.write(f"    {d.mapping_rule} AS {d.column_name}")
            if i < len(self.data) - 1:
                w.write(",")
            w.write(f"    -- {d.column_cn_name}")
            w.write("\n")

        mt = self.main_table()
        w.write(f"FROM {mt.source_table} {mt.source_table_id}\n")
        for d in filter(lambda x: x.is_join(), self.data):
            w.write(
                f"{d.join_type} {d.source_table} {d.source_table_id} {d.join_on_condition}\n"
            )
        w.write(f"{mt.where_condition};\n")

    def __str__(self) -> str:
        sio = StringIO()
        self.write_sql(sio)
        return sio.getvalue()


class ModelFile:
    def __init__(self, sheet: Worksheet) -> None:
        self.sheet_name = sheet.title
        self.meta = []
        for row in range(2, 12):
            self.meta.append((sheet.cell(row, 1).value, sheet.cell(row, 2).value))

        groups = {}

        for v in sheet.iter_rows(min_row=14, values_only=True):
            if v[0] is not None:
                gid = v[5]
                if gid not in groups:
                    groups[gid] = DataGroup(gid)
                group: DataGroup = groups[gid]
                group.data.append(DataRow(v))

        self.data: list[DataGroup] = list(groups.values())

    @property
    def target_table_name(self) -> str:
        return self.meta[2][1]

    @property
    def target_table_cn_name(self) -> str:
        return self.meta[1][1]

    def table_info(self):
        array = self.target_table_name.split(".")
        if len(array) == 1:
            return None, self.target_table_name
        return array[0], array[1]

    def write_meta_comments(self, writer: EnhancedWriter):
        for kv in self.meta:
            key = str(kv[0]).replace("\n", " ")
            vs = str(kv[1]).split("\n")
            writer.write(f"-- {key.ljust(10,chr(12288))}: {vs[0]}\n")
            if len(vs) > 1:
                for v in vs[1:]:
                    writer.write(f"-- {''.ljust(10,chr(12288))}  {v}\n")

    def write_content(self, writer: EnhancedWriter):
        self.write_meta_comments(writer)
        writer.writeln()
        self.write_postgre_procedure(writer)

    def write_postgre_procedure(self, writer: EnhancedWriter):
        info = self.table_info()
        function_name = "dcx_" + info[1]
        if info[0] is not None:
            function_name = f"{info[0]}.dcx_{info[1]}"
        writer.writeln(
            f"CREATE OR REPLACE FUNCTION {function_name}(IN p_date int4, OUT p_srccnt int4, OUT p_dstcnt int4) AS $$"
        )
        writer.writeln("BEGIN")
        writer.writeln()

        temp_table_name = "tmp_" + info[1]
        writer.writeln(f"DROP TABLE IF EXISTS {temp_table_name};")
        writer.writeln(
            f"CREATE TEMPORARY TABLE {temp_table_name} AS (SELECT * FROM {self.target_table_name} WHERE 1=2);"
        )
        for group in self.data:
            if group.group_id is None:
                continue
            writer.writeln(f"-- Group {group.group_id}")
            writer.writeln(f"INSERT INTO {temp_table_name}")
            writer.writeln(group, 4)

        writer.writeln(f"DELETE FROM {self.target_table_name} WHERE dcdate=p_date;")
        writer.writeln(
            f"INSERT INTO {self.target_table_name} SELECT * FROM {temp_table_name};"
        )
        writer.writeln()
        writer.writeln(f"SELECT count(1) INTO p_srccnt FROM {temp_table_name};")
        writer.writeln("p_dstcnt := p_srccnt;")
        writer.writeln()
        writer.writeln("END;")
        writer.writeln("$$ LANGUAGE plpgsql;")

    def generate_etl_file(self, args: Args):
        file = os.path.join(args.dest, f"{self.target_table_name}.sql")
        with EnhancedWriter(file) as writer:
            self.write_content(writer)
        log.info(f"生成ETL脚本：{self.sheet_name} > {file}")

    def generate_ddl_file(self, args: Args):
        file = os.path.join(args.dest, f"{self.target_table_name}.ddl.sql")
        with EnhancedWriter(file) as writer:
            self.write_postgre_ddl(writer)
        log.info(f"生成DDL脚本：{self.sheet_name} > {file}")

    def get_primary_keys(self):
        group = self.data[0]
        keys = filter(lambda x: x.column_is_pk, group.data)
        return list(map(lambda x: x.column_name, keys))

    def write_postgre_ddl(self, w: EnhancedWriter):
        w.writeln(f"-- Table: {self.target_table_name} {self.target_table_cn_name}")
        pks = self.get_primary_keys()
        w.writeln(f"CREATE TABLE {self.target_table_name} (")
        group = self.data[0]
        for i in range(len(group.data)):
            d = group.data[i]
            w.write(f"    {d.column_name} {d.column_type} {d.column_null_str}")
            # if d.column_default_value is not None:
            #     w.write(f" DEFAULT {d.column_default_value}")
            if i < len(group.data) - 1 or len(pks) > 0:
                w.write(",")
            w.writeln(f"    -- {d.column_cn_name}")
        if len(pks) > 0:
            info = self.table_info()
            w.writeln(f"    CONSTRAINT pk_{info[1]} PRIMARY KEY ({','.join(pks)})")

        w.writeln(f");")

        w.writeln(
            f"COMMENT ON TABLE {self.target_table_name} IS '{self.target_table_cn_name}';"
        )
        for i in range(len(group.data)):
            d = group.data[i]
            w.writeln(
                f"COMMENT ON COLUMN {self.target_table_name}.{d.column_name} IS '{d.column_cn_name}';"
            )


def find_model_file():
    cwd = os.getcwd()
    entrys = os.listdir(cwd)

    for entry in entrys:
        file = os.path.join(cwd, entry)
        if (
            os.path.isfile(file)
            and file.endswith(".xlsx")
            and not entry.startswith("~$")
        ):
            return file
    return None


def is_model_sheet(sheet: Worksheet):
    if sheet.max_row < 13:
        return False
    A13: ReadOnlyCell = sheet["A13"]
    if A13.value != "目标字段英文名":
        return False
    return True


def load_model_file(args: Args):

    book: Workbook = load_workbook(
        args.file, read_only=True, data_only=True, keep_vba=False
    )

    for sheet_name in book.sheetnames:
        sheet: Worksheet = book[sheet_name]
        if is_model_sheet(sheet):
            model = ModelFile(sheet)
            yield model


def main():

    parser = argparse.ArgumentParser(description="ETL脚本生成程序")
    parser.add_argument(
        "--version", action="version", version=__version__, help="显示程序版本号"
    )
    parser.add_argument(
        "file",
        type=str,
        nargs="?",
        default=None,
        help="Excel模型文件(*.xlsx)，如果不指定会从当前工作目录自动寻找第一个模型文件进行处理。",
    )
    parser.add_argument("--dest", default=".", type=str, help="设定文件生成目录，默认为当前工作目录。")

    args: Args = parser.parse_args()

    log.info(f"{parser.description} {__version__}")

    if args.file is None:
        args.file = find_model_file()

    if args.file is None:
        log.error("未找到Excel模型文件")
        exit(1)

    os.makedirs(args.dest, exist_ok=True)

    for model in load_model_file(args):
        model.generate_etl_file(args)
        model.generate_ddl_file(args)


if __name__ == "__main__":
    main()
