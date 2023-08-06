import pandas as pd
from openpyxl import load_workbook

import openpyxl
import xlrd

def check_for_charts_graphs_images(sheet_name):
  workbook = openpyxl.load_workbook('171testcosmic.xls')
  sheet = workbook[sheet_name]
  print(sheet.tables)
  print(list(sheet.values))
  for t in list(sheet.values):
      for v in t:
          if v is not None:
              print(v)
  for row in sheet.iter_rows():
      for cell in row:
          if cell.value is not None:
              print(cell)
              print(cell.value)
  return False

# print(check_for_charts_graphs_images(sheet_name='系数表'))

def test():
    df = pd.read_excel('171testcosmic.xls', sheet_name='结算评估确认表', engine='xlrd', skiprows=(0,))

    from tabulate import tabulate
    print(tabulate(df))

    print(df)


test()

def st():
    df = pd.read_excel('st.xlsx')

    print(df)
    # print(df.loc['A', 'E'])

st()

import re

pattern = re.compile("\/(?P<req>[0-9]{1,4})\/faqi\/[0-9A-Za-z,&@#$%.\[\]{};']+\.xlsx$")

if pattern.match("/454/faqi/hello.xlsx"):
    print(pattern.groups)
    print(pattern.groupindex)
    print(pattern.flags)

for match in re.finditer(pattern, '/123/faqi/asida.xlsx'):
    print(match)
    print(match['req'])
    print(type(match['req']))

df = pd.read_excel('171testcosmic.xls', sheet_name='系数表', engine='xlrd')
num_s : pd.Series = df['数值']
print("last", num_s.iloc[-1])
print("last df", df.iloc[-1, -1])

# pd.read_excel("st.xlsx", sheet_name="asjdoasjodajosdj")  # ValueError
df = pd.read_excel('st.xlsx')
print(df)
print('-' * 50)
df.columns = df.iloc[0]
print(df)
print('-' * 50)
df = df.iloc[1:]
print(df)
print('-' * 50)
df = df.reset_index(drop=True)
print(df)
print(df.iloc[1, 2])

import xlrd
book = xlrd.open_workbook('st.xls', formatting_info=True)
sheet = book.sheet_by_name('Sheet1')
print(sheet)

rows = sheet.nrows
cols = sheet.ncols

for i in range(rows):
    for j in range(cols):
        # print('val', sheet.cell_value(rowx=i, colx=j))
        cell = sheet.cell(rowx=i, colx=j)
        xfx = sheet.cell_xf_index(rowx=i, colx=j)  # xf index of given cell, index into xf_list
        xf = book.xf_list[xfx]  # a list of xf class instances, each corresponding to an XF record
        bgx = xf.background.pattern_colour_index
        # print(bgx)

        print(i, j, cell.value, bgx)

        # if bgx == 13:
        #     print(i, j, cell.value, 'yellow')

'''
    xlrd xf
    0x08 - 0x0a: 8 fixed colour xf index
    0x18, 0x40: SYS WIN text colour for border lines
    0x19, 0x41: bg 
    
    standard color: num (decimal)
    dark red: 60
    red: 10
    orange: 51
    yellow: 13
    light green: 50
    green: 17
    light blue: 40
    blue: 30
    dark blue: 56
    purple: 36
    white: 9
    no fill: 64
'''

book = load_workbook('st.xlsx')
sheet = book['Sheet1']
print(type(sheet.columns))
print(sheet.columns)
for row in sheet.iter_rows():
    print(row, type(row))
    for cell in row:
        print(cell)
        color_in_hex : str = cell.fill.start_color.index
        print(cell.value, color_in_hex, type(color_in_hex))
        print(cell.row, cell.column)  # get index of the cell, 1-based

for i in range(len(list(sheet.rows))):
    row = list(sheet.rows)[i]
    print(row)

    for j in range(len(row)):
        cell = row[j]
        print(i, j, cell)

from deprecated import deprecated

@deprecated(version='0.1.0', reason='deprecated in 0.2.0')
def hello():
    print("hello")

hello()

a = float('0.33')
b = 1/3
if abs(a - b) < 0.01:
    print('e')
else:
    print('ne')








