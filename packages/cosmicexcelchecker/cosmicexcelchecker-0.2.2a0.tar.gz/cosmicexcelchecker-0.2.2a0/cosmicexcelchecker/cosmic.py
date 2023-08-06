# Core COSMIC File

from _baseclass import PdExcel
from typing import Union, Dict, List
from errors import CosmicExcelCheckerException ,IncorrectFileTypeException, RepeatedREQNumException, \
    SheetNotFoundException, UnknownREQNumException
from tabulate import tabulate
from conf import CFP_SHEET_NAMES ,CFP_COLUMN_NAME, SUB_PROCESS_NAME, RS_SKIP_ROWS, RS_TOTAL_CFP_NAME, \
    RS_WORKLOAD_NAME, RS_REQ_NUM, RS_REQ_NAME, SR_COSMIC_REQ_NAME,  SR_NONCOSMIC_REQ_NAME, SR_SUBFOLDER_NAME, \
    SR_COSMIC_FILE_PREFIX, SR_NONCOSMIC_FILE_PREFIX, RS_QLF_COSMIC, COEFFICIENT_SHEET_NAME, \
    COEFFICIENT_SHEET_DATA_COL_NAME, NONCFP_SHEET_NAMES, SR_NONCOSMIC_PROJECT_NAME, SR_NONCOSMIC_REQ_NUM, \
    SR_AC_REPORT_NUM, SR_AC_FINAL_NUM, SR_FINAL_CONFIRMATION, SR_AC_REQ_NUM, SR_AC_REQ_NAME, SR_AC_FINAL_NUM_LIMIT

from find import FindExcels

from deprecated import deprecated

import pandas as pd
import numpy as np
import time
import re
import math
import xlrd
import openpyxl

class CosmicReqExcel(PdExcel):
    '''
    Implementation of abstract class PdExcel
    Read a single Excel/CSV about Cosmic Requirement
    '''

    def __init__(self, path: str):
        '''
        data_frames is the pd.DataFrame converted from Spreadsheet
        log holds temporary error/warning for later usage (e.g print to terminal)

        :param path: path to single requirement file
        '''
        self.path : str = path
        self.data_frames: Union[Dict[str, pd.DataFrame], None] = None
        self.log : Union[List[str], str, None] = None
        self.file_format : Union[str, None] = None

    def load_excel(self):
        file_ext = self.path[self.path.rindex('.'):]

        # it can be simplified to a dict[file_ext:engine] but with less readability
        if file_ext in ('.xlsx', '.xls'):
            self.data_frames = pd.read_excel(self.path, sheet_name=None)

            self.file_format = file_ext
        else:
            raise IncorrectFileTypeException(f"{self.path} is not a valid relative file path for an Excel file")

    def print_df(self):
        '''
        Try to print the converted pd.Dataframe to the terminal

        :return: None
        '''

        for df in self.data_frames.values() if isinstance(self.data_frames, dict) else self.data_frames:

            print(tabulate(df, headers='keys', tablefmt='psql'))

    def get_req_name(self) -> Union[str, None]:
        '''
        Get the requirement name of a single req file

        :return: name string or None if not no data found
        '''

        cfp_df : Union[pd.DataFrame, None] = None
        for sheet_name in CFP_SHEET_NAMES:  # iterate through
            cfp_df = self.data_frames.get(sheet_name, None) if isinstance(self.data_frames, dict) else None

            if cfp_df is not None:
                break

        if cfp_df is None:  # noqa
            return None

        # get column name by iterating through .columns
        for col_name in cfp_df.columns:
            if col_name.startswith(SR_COSMIC_REQ_NAME):
                return cfp_df.iloc[0, cfp_df.columns.get_loc(col_name)]

        return None


    def get_CFP_total(self) -> Union[float, None]:
        '''
        get total CFP pts under CFP column
        it will convert all possible numeric values to dtype float (or int) and leave others as NaN

        :return: total CFP pts
        '''
        data : Union[pd.DataFrame, None] = None
        for sheet_name in CFP_SHEET_NAMES:  # iterate through
            data = self.data_frames.get(sheet_name, None) if isinstance(self.data_frames, dict) else None

            if data is not None:
                break

        if data is None or CFP_COLUMN_NAME not in data.columns or SUB_PROCESS_NAME not in data.columns:
            return None

        cfp_s : pd.Series = pd.to_numeric(data.loc[:, CFP_COLUMN_NAME], errors='coerce')  # convert to numeric
        cfp_df : pd.DataFrame = data.loc[:, [SUB_PROCESS_NAME, CFP_COLUMN_NAME]]

        sub_process_null_count = 0  # only when subprocess is null, not both null
        for index, row in cfp_df.iterrows():
            if pd.isna(row[SUB_PROCESS_NAME]) and not pd.isna(row[CFP_COLUMN_NAME]):
                sub_process_null_count += 1

        return cfp_s.sum() - sub_process_null_count

    def check_CFP_column(self) -> dict:
        '''
        check cfp column with incorrect value
        Correct values: [0, 1]

        :return: a dict-format result
        '''

        for sheet_name in CFP_SHEET_NAMES:  # iterate through
            cfp_df : Union[pd.DataFrame, None] = self.data_frames.get(sheet_name, None) if isinstance(self.data_frames, dict) else None

            if cfp_df is not None:
                break

        if cfp_df is None or CFP_COLUMN_NAME not in cfp_df or SUB_PROCESS_NAME not in cfp_df:  # noqa
            return {'path': self.path, 'match': False, 'CFP': -1, 'note': 'No CFP related column'}

        cfp_df : pd.DataFrame = cfp_df.loc[:, [SUB_PROCESS_NAME, CFP_COLUMN_NAME]]  # noqa
        cfp_total = self.get_CFP_total()  # it will always return a float since CFP column exists

        # record if cfp or sub-process column miss anything
        note = f"Missing data in {'CFP Column' * int(cfp_df.loc[:, CFP_COLUMN_NAME].isna().sum() > 0)} " \
               f"{'Subprocess description' * int(cfp_df.loc[:, SUB_PROCESS_NAME].isna().sum() > 0)}".strip()
        # if cfp_df.loc[:, CFP_COLUMN_NAME].isna().sum() > 0:  # blank space or non-numeric char
        #     note = 'Missing data in CFP Column'
        #
        # if cfp_df.loc[:, SUB_PROCESS_NAME].isna().sum() > 0:
        #     note +=
        if note == 'Missing data in':  # meaning there's no missing
            note = ''

        return {
            "path": self.path,
            "match": True,
            "CFP": cfp_total,
            "note": note
        }

    def check_coefficient_sheet(self) -> Union[bool, None]:
        '''
        check CFP pts is the same with the B1 CFP pts in coefficient sheet if applicable

        :return: match/not match (bool) or None if there's no coefficient_sheet
        '''

        coefficient_sheet: pd.DataFrame = self.data_frames.get(COEFFICIENT_SHEET_NAME, None)

        if coefficient_sheet is None:
            return None

        # print(coefficient_sheet)
        std_cfp_pts = coefficient_sheet.iloc[1, coefficient_sheet.columns.get_loc(COEFFICIENT_SHEET_DATA_COL_NAME)]

        if std_cfp_pts is None or std_cfp_pts == '':
            return None

        try:
            std_cfp_pts = float(std_cfp_pts)

            return std_cfp_pts == self.get_CFP_total()

        except ValueError:
            return None

    def check_final_confirmation(self) -> dict:
        '''
        check final confirmation worksheet if applicable, by the contents comparing to itself and other sheets

        :return: a result dictionary shows the match info
        '''
        for sheet_name in SR_FINAL_CONFIRMATION:
            fc_sheet : Union[pd.DataFrame, None] = self.data_frames.get(sheet_name, None) if \
                isinstance(self.data_frames, dict) else None

            if fc_sheet is not None:
                break

        if fc_sheet is None:  # noqa
            return {
                "path": self.path,
                "match": False,
                "note": "No related final confirmation worksheet found"
            }

        # ignore first row, set index 1 as column and reset index.
        fc_sheet.columns = fc_sheet.iloc[0]  # noqa
        fc_sheet = fc_sheet.iloc[1:]
        fc_sheet = fc_sheet.reset_index(drop=True)

        # load other sheets for comparison
        coefficient_sheet : pd.DataFrame = self.data_frames.get(COEFFICIENT_SHEET_NAME, None) \
            if isinstance(self.data_frames, dict) else None

        if coefficient_sheet is None:
            return {
                "path": self.path,
                "match": False,
                "note": "No related coeffficient worksheet found"
            }

        for sheet_name in CFP_SHEET_NAMES:  # iterate through
            cfp_df : Union[pd.DataFrame, None] = self.data_frames.get(sheet_name, None) \
                if isinstance(self.data_frames, dict) else None

        if cfp_df is None:  # noqa
            return {
                "path": self.path,
                "match": False,
                "note": "No related Cosmic worksheet found"
            }

        try:
            note : str = ''

            # extract folder req by path
            extract_folder_req = rf"\/(?P<req>[0-9]{{1,4}})\/{SR_SUBFOLDER_NAME}\/{SR_COSMIC_FILE_PREFIX}[0-9A-Za-z,&@#$%.\[\]{{}};'\u4e00-\u9fff：。，（）()’￥……]+\.xls[x]{{0,1}}$"
            match = ""
            for m in re.finditer(pattern=extract_folder_req, string=self.path):
                match = m['req']  # type: str

            if match == '':
                note += 'Path Corrupted (no req num subfolder)\t'

            if not match.isnumeric():
                note += 'Missing Req num (cosmic, fc)'

            elif fc_sheet.iloc[0, fc_sheet.columns.get_loc(SR_AC_REQ_NUM)] != int(match):
                    note += 'Req Num not match (cosmic, fc)\t'

            # check req name
            if fc_sheet.iloc[0, fc_sheet.columns.get_loc(SR_AC_REQ_NAME)] != self.get_req_name():
                note += 'Req Name not match (cosmic, fc)\t'

            # **DO NOT DELETE**
            # # check report num of days
            # std_num_days = round(coefficient_sheet.iloc[-1, -1])
            # if abs(fc_sheet.iloc[0, fc_sheet.columns.get_loc(SR_AC_REPORT_NUM)] - std_num_days) >= 0.1:  # errors
            #     note += 'report num of days not match (coefficient, fc)\t'
            #
            # # check final report num of days
            # if std_num_days <= SR_AC_FINAL_NUM_LIMIT:
            #     final_num_days = std_num_days
            # else:
            #     final_num_days = SR_AC_FINAL_NUM_LIMIT
            #
            # fnd_inchart = fc_sheet.iloc[0, fc_sheet.columns.get_loc(SR_AC_FINAL_NUM)]
            # if fnd_inchart > SR_AC_FINAL_NUM_LIMIT or abs(fnd_inchart - final_num_days) >= 0.1:  # errors
            #     note += 'final num of days not match (fc)\t'
            #
            # note = note.strip('\t')

            return {
                "path": self.path,
                "match": note == '',
                "note": note
            }

        except KeyError:
            return {
                "path": self.path,
                "match": False,
                "note": "Key Error in worksheet. Make sure they are in standard format"
            }

    def check_highlight_cfp(self) -> list[int, None]:
        '''
        check the highlight on sub-process and its corresponding cfp in the same line
        No fill: 1 cfp; Yellow: 0 cfp; Red: 1/3 cfp

        :return:
        '''

        # load the Excel again using openpyxl/xlrd
        if self.file_format == '.csv':
            return list()
        elif self.file_format == '.xlsx':
            excel = openpyxl.load_workbook(self.path)

            # open cfp sheet
            sheet = None
            cfp_df : Union[pd.DataFrame, None] = None
            for sheet_name in CFP_SHEET_NAMES:
                if sheet_name in excel:
                    sheet = excel[sheet_name]
                    cfp_df = self.data_frames.get(sheet_name, None) if \
                        isinstance(self.data_frames, dict) else None
                    break

            if sheet is None or cfp_df is None:
                raise SheetNotFoundException(f"Standard CFP Sheet not found")

            # get subprocess index and cfp
            sp_idx = cfp_df.columns.get_loc(SUB_PROCESS_NAME)  # subprocess col
            cfp_idx = cfp_df.columns.get_loc(CFP_COLUMN_NAME)

            err_list : list = []
            row_num = 1  # 1-based, =1 since min_row = 2
            # extract cell value and compare
            for row in sheet.iter_rows(min_row=2, max_row=len(cfp_df.index) + 1, min_col=sp_idx + 1, max_col=sp_idx + 1):  # 1-based idx for min_row, each row is a tuple
                row_num += 1
                sp_color_hex : str = ""
                cfp_cell : str = ""
                # for i in range(len(row)):
                #     if i == sp_idx and row[i].value != SUB_PROCESS_NAME:  # avoid first row
                #         sp_color_hex : str = row[i].fill.start_color.index  # hex value for the cell fill color
                #
                #     if i == cfp_idx and row[i].value != CFP_COLUMN_NAME:
                #         cfp_cell = str(row[i].value)
                sp_color_hex : str = row[0].fill.start_color.index
                cfp_cell = str(cfp_df.iloc[row_num - 2, cfp_idx])  # avoid str cell value

                if sp_color_hex == "":  # not counted since subprocess is empty
                    continue

                if cfp_cell == "":  # only count valid subprocess row
                    err_list.append(f'{row_num} Missing Data')
                    continue
                try:
                    cfp_cell : float = float(cfp_cell)
                except ValueError:
                    err_list.append(f'{row_num} CFP not a number')
                    continue

                if sp_color_hex == 'FFFFFF00' and cfp_cell != 0:  # YELLOW
                    err_list.append(f'{row_num} Yellow != 0')
                elif sp_color_hex == 'FFFF0000' and abs(cfp_cell - 1/3) >= 0.01:  # RED
                    err_list.append(f'{row_num} Red != 1/3 or 0.333')
                elif (sp_color_hex == '00000000' or (type(sp_color_hex) is int and sp_color_hex == 9)) and cfp_cell != 1:  # No fill
                    err_list.append(f'{row_num} No fill (White) != 1')

        elif self.file_format == '.xls':
            excel = xlrd.open_workbook(self.path, formatting_info=True)

            # open cfp sheet
            sheet = None
            cfp_df: Union[pd.DataFrame, None] = None
            for sheet_name in CFP_SHEET_NAMES:
                if sheet_name in excel.sheet_names():
                    sheet = excel.sheet_by_name(sheet_name)
                    cfp_df = self.data_frames.get(sheet_name, None) if \
                        isinstance(self.data_frames, dict) else None
                    break

            if sheet is None or cfp_df is None:
                raise SheetNotFoundException(f"Standard CFP Sheet not found")

            # get idx of subprocess and cfp col
            sp_idx = cfp_df.columns.get_loc(SUB_PROCESS_NAME)  # subprocess col
            cfp_idx = cfp_df.columns.get_loc(CFP_COLUMN_NAME)

            err_list = []
            rows = min(sheet.nrows, len(cfp_idx.index) + 1)

            for i in range(1, rows):
                sp_cell : xlrd.sheet.Cell = sheet.cell(rowx=i, colx=sp_idx)
                # cfp_cell : xlrd.sheet.Cell = sheet.cell(rowx=i, colx=cfp_idx)

                xfx = sheet.cell_xf_index(rowx=i, colx=sp_idx)  # xf index
                xf = excel.xf_list[xfx]  # used xfx as index
                bgx = xf.background.pattern_colour_index

                cfp : str = str(cfp_df.iloc[i - 1, cfp_idx])
                if str(sp_cell.value) == "":  # not count as valid if subprocess is empty
                    continue

                if cfp == "":  # only count valid subprocess row
                    err_list.append(f'{i + 1} Missing data')
                    continue

                try:
                    cfp : float = float(cfp)
                except ValueError:
                    err_list.append(f'{i + 1} CFP not a number')
                    continue

                if bgx == 13 and float(cfp) != 0:  # YELLOW
                    err_list.append(f'{i + 1} Yellow != 0')
                elif bgx == 10 and abs(float(cfp) - 1/3) >= 0.01:  # RED
                    err_list.append(f'{i + 1} Red != 1/3 or 0.333')
                elif bgx == 64 and float(cfp) != 1:  # no fill
                    err_list.append(f'{i + 1} No fill (White) != 1')

        else:
            raise IncorrectFileTypeException(f"Incorrect file type {self.file_format}. It has to be .xlsx or .xls file (.csv deprecated)")

        return err_list

class NonCosmicReqExcel(PdExcel):
    '''
    Implementation of Abstract class PdExcel
    Mainly used for representing non-cosmic Excel file (single requirement)
    '''

    def __init__(self, path: str):
        '''
                data_frames is the pd.DataFrame converted from Spreadsheet
                log holds temporary error/warning for later usage (e.g print to terminal)

                :param path: path to single requirement file
                '''
        self.path: str = path
        self.data_frames: Union[Dict[str, pd.DataFrame], None] = None
        self.log: Union[List[str], str, None] = None

    def load_excel(self):
        file_ext = self.path[self.path.rindex('.'):]

        # it can be simplified to a dict[file_ext:engine] but with less readability
        if file_ext in ('.xlsx', '.xls'):
            self.data_frames = pd.read_excel(self.path, sheet_name=None)
        else:
            raise IncorrectFileTypeException(f"{self.path} is not a valid relative file path for an Excel file")

    def print_df(self):
        '''
        Try to print the converted pd.Dataframe to the terminal

        :return: None
        '''

        for df in self.data_frames.values() if isinstance(self.data_frames, dict) else self.data_frames:
            print(tabulate(df, headers='keys', tablefmt='psql'))

    def get_req_name(self) -> Union[str, None]:
        '''
        Get requirement name for the Excel

        :return: requirement as a string or None if no related column found
        '''

        workload_df = self.data_frames.get(NONCFP_SHEET_NAMES, None)

        if workload_df is None or SR_NONCOSMIC_REQ_NAME not in workload_df.columns:
            return None

        try:
            return workload_df.iloc[0, workload_df.columns.get_loc(SR_NONCOSMIC_REQ_NAME)]
        except IndexError:  # if not exists
            return None

    def get_project_name(self) -> Union[str, None]:
        '''
        Get project name for the Excel

        :return: requirement as a string or None if no related column found
        '''

        workload_df = self.data_frames.get(NONCFP_SHEET_NAMES, None)

        if workload_df is None or SR_NONCOSMIC_PROJECT_NAME not in workload_df.columns:
            return None

        try:
            return workload_df.iloc[0, workload_df.columns.get_loc(SR_NONCOSMIC_PROJECT_NAME)]
        except IndexError:  # if not exists
            return None

class ResultSummary(PdExcel):
    '''
    Another implementation of Abstract class PdExcel
    Demonstrated for loading and processing data in Result Summary related excels
    '''

    def __init__(self, path: str, folders_path:str, sheet_name: str):
        '''
        data_frames is the pd.DataFrame converted from Spreadsheet
        log holds temporary error/warning for later usage (e.g print to terminal)

        :param path: path of the result summary file
        :param sheet_name: specific worksheet name in result summary to be loaded
        '''
        self.path : str = FindExcels.path_format(path=path)
        self.folders_path : str = FindExcels.path_format(path=folders_path)
        self.data_frames: Union[Dict[str, pd.DataFrame], None] = None
        self.data_frame_specific : Union[pd.DataFrame, None] = None
        self.log : Union[list[str], str, None] = None
        self.sheet_name : str = sheet_name
        self.file_paths : Union[list[str, None], None] = FindExcels.find_excels(path=self.folders_path)

    def load_excel(self):
        file_ext = self.path[self.path.rindex('.'):]

        # it can be simplified to a dict[file_ext:engine] but with less readability
        if file_ext in ('.xlsx', '.xls'):
            self.data_frames = pd.read_excel(self.path, sheet_name=None, skiprows=range(RS_SKIP_ROWS))

            df_specific = self.data_frames.get(self.sheet_name, None)

            if df_specific is None:
                raise SheetNotFoundException(f"Sheet with name {self.sheet_name} is not found inside the given file")

            self.data_frame_specific = df_specific
        else:
            raise IncorrectFileTypeException(f"{self.path} is not a valid relative file path for an Excel file")

    def set_sheet_name(self, sheet_name: str):
        '''
        Setting worksheet name for reading dataframe later

        :param sheet_name: str, name of the worksheet
        :return: None
        '''

        self.sheet_name = sheet_name

        df_specific = self.data_frames.get(self.sheet_name, None)

        if df_specific is None:
            raise SheetNotFoundException(f"Sheet with name {sheet_name} is not found inside the given file")

        self.data_frame_specific = df_specific


    def print_df(self):
        '''
        Try to print the converted pd.Dataframe to the terminal

        :return: None
        '''

        for df in self.data_frames.values() if isinstance(self.data_frames, dict) else self.data_frames:

            print(tabulate(df, headers='keys', tablefmt='psql'))


    def print_df_specific(self):
        '''
        Try to print the specific sheet set by function set_sheet_name
        :return: None
        '''

        if isinstance(self.data_frame_specific, pd.DataFrame):
            print(tabulate(self.data_frame_specific, headers='keys', tablefmt='psql'))

        else:
            raise CosmicExcelCheckerException("Specific worksheet is not loaded. Use `set_sheet_name` to load it")

    def check_ratio(self) -> list[str, None]:
        '''
        check all columns workload and cfp ratio, default to 0.79

        :return: a list contains all non-qualified requirements
        '''

        if self.data_frame_specific is None or RS_WORKLOAD_NAME not in self.data_frame_specific.columns or RS_TOTAL_CFP_NAME not in self.data_frame_specific.columns:
            return list()

        result_summary = self.data_frame_specific.loc[:, [RS_REQ_NUM, RS_REQ_NAME, RS_WORKLOAD_NAME,RS_TOTAL_CFP_NAME]]

        bad_ratio : list[str, None] = []  # record
        for index, row in result_summary.iterrows():
            if math.ceil(int(row[RS_WORKLOAD_NAME]) / 0.79) < int(row[RS_TOTAL_CFP_NAME]):
                bad_ratio.append(
                    f'{RS_REQ_NUM}{row[RS_REQ_NUM]}, {RS_REQ_NAME}: {row[RS_REQ_NAME]}'
                )

        return bad_ratio

    def check_file(self, req_num: int, check_final_confirmation: bool = True,
                   check_highlight_cfp: bool = True) -> dict:
        '''
        check a single file data, comparing to the result summary xlsx
        check req number, req name, CFP total, CFP Total comparison

        :param: req_num is the requirement number 需求序号
        :return: a dict-format result
        '''

        try:
            if self.data_frame_specific is None:
                raise CosmicExcelCheckerException()
        except CosmicExcelCheckerException:
            print("Specific worksheet is not loaded. Use `set_sheet_name` to load it")
            return dict()

        # get row indices qualified for req_num
        r_indices : list = self.data_frame_specific.index[self.data_frame_specific[RS_REQ_NUM] == req_num].tolist()

        try:
            if len(r_indices) < 0:
                raise UnknownREQNumException()

            elif len(r_indices) > 1:
                raise RepeatedREQNumException()

        except UnknownREQNumException:
            print(f"Sheet does not have a requirement number called {req_num}")
            return dict()

        except RepeatedREQNumException:
            print(f"Sheet has repeated rows for requirement number {req_num}")
            return dict()

        row_index = r_indices[0]

        req_num = self.data_frame_specific.iloc[row_index, self.data_frame_specific.columns.get_loc(RS_REQ_NUM)]

        # requirement folder path based on req_num
        req_folder_path_pattern = rf'{self.folders_path}\/(?:.*\/|){str(req_num)}\/'

        # qualified paths inside self.file_paths
        qualified_paths : list = [path for path in self.file_paths if re.match(pattern=req_folder_path_pattern, string=path)]

        if len(qualified_paths) == 0:  # no subfolder found
            return {"REQ Num": req_num, "path": "Not exist", "match": False, "note": "REQ folder does not exist"}

        qualified_cosmic = self.data_frame_specific.iloc[row_index, self.data_frame_specific.columns.get_loc(RS_QLF_COSMIC)]

        # check sr cosmic
        def check_cosmic(path: str):
            cosmic_excel = CosmicReqExcel(path=path)

            # load excel to class df
            cosmic_excel.load_excel()

            note = ''
            # check req name
            if self.data_frame_specific.iloc[
                row_index, self.data_frame_specific.columns.get_loc(RS_REQ_NAME)] != cosmic_excel.get_req_name():
                note += 'REQ name does not match (cosmic); '

            # check total CFP name
            total_cfp: str = str(
                self.data_frame_specific.iloc[row_index, self.data_frame_specific.columns.get_loc(RS_TOTAL_CFP_NAME)])
            if total_cfp.isnumeric():
                if float(total_cfp) != cosmic_excel.get_CFP_total():
                    note += 'Total CFP points do not match (cosmic); '
            else:
                note += 'CFP points in Result Summary is not valid (cosmic); '

            # Check coefficient sheet
            coefficient_sheet_match = cosmic_excel.check_coefficient_sheet()

            if coefficient_sheet_match is None:
                note += 'No Coefficient Sheet in Excel (cosmic); '
            elif coefficient_sheet_match is False:
                note += 'Coefficient Sheet B1 data does not match total standard CFP pts; '

            # check final confirmation worksheet
            if check_final_confirmation:
                fc_result = cosmic_excel.check_final_confirmation()
                if fc_result['note'] != "":
                    note += f"{fc_result['note']}; "

            # check highlight
            if check_highlight_cfp:
                hl_result = cosmic_excel.check_highlight_cfp()
                if hl_result != list():
                    note += f"highlight err: {str(hl_result)}; "

            note = note.rstrip('; ')

            return {"REQ Num": req_num, "path": path, "match": note == "", "note": note}

        # check sr noncosmic
        def check_noncosmic(path: str):
            noncosmic_excel = NonCosmicReqExcel(path=path)

            # load excel to class df
            noncosmic_excel.load_excel()

            note = ''
            # check req name
            if self.data_frame_specific.iloc[
                row_index, self.data_frame_specific.columns.get_loc(RS_REQ_NAME)] != noncosmic_excel.get_req_name():
                note += 'REQ name does not match (noncosmic); '

            # make sure cfp total is 0 for non-cosmic file
            # total_cfp: str = str(
            #     self.data_frame_specific.iloc[row_index, self.data_frame_specific.columns.get_loc(RS_TOTAL_CFP_NAME)])
            # if float(total_cfp) != cosmic_cfp:  # cosmic_cfp is total cfp checked by cosmic file
            #     note += 'Total CFP points is not 0 for non-cosmic requirement; '

            note = note.rstrip('; ')

            return {"REQ Num": req_num, "path": path, "match": note == "", "note": note}

        # qualified paths parent folder
        qualified_paths_docs = [path[path.rindex('/') + 1:] for path in qualified_paths]

        if qualified_cosmic == '是':
            if len(qualified_paths) == 1:

                if not qualified_paths_docs[0].startswith(SR_COSMIC_FILE_PREFIX):
                    return {"REQ Num": req_num, "path": qualified_paths[0], "match": False, "note": "Incorrect type of cosmic excel based on requirement"}

                # file matched
                try:
                    return check_cosmic(path=qualified_paths[0])
                except KeyError:
                    return {"REQ Num": req_num, "path": qualified_paths[0], "match": False, "note": "KeyError, Check column name"}

            else:
                return {"REQ Num": req_num, "path": qualified_paths[0], "match": False, "note": "Incorrect number of cosmic excel(s) based on requirement"}

        elif qualified_cosmic == '否':
            if len(qualified_paths) == 1:

                if not qualified_paths_docs[0].startswith(SR_NONCOSMIC_FILE_PREFIX):
                    return {"REQ Num": req_num, "path": qualified_paths[0], "match": False,
                            "note": "Incorrect type of cosmic excel based on requirement"}

                # file matched
                try:
                    return check_noncosmic(path=qualified_paths[0])
                except KeyError:
                    return {"REQ Num": req_num, "path": qualified_paths[0], "match": False,
                            "note": "KeyError, Check column name"}

            else:
                return {"REQ Num": req_num, "path": qualified_paths[0], "match": False,
                        "note": "Incorrect number of non-cosmic excel(s) based on requirement"}

        elif qualified_cosmic == '混合型':
            if len(qualified_paths) == 2:
                c_prefix_pattern = rf"{req_folder_path_pattern}(?:.*\/|){SR_SUBFOLDER_NAME}\/{SR_COSMIC_FILE_PREFIX}"
                nc_prefix_pattern = rf"{req_folder_path_pattern}(?:.*\/|){SR_SUBFOLDER_NAME}\/{SR_NONCOSMIC_FILE_PREFIX}"

                try:
                    if bool(re.match(c_prefix_pattern, qualified_paths[1])) and \
                            bool(re.match(nc_prefix_pattern, qualified_paths[0])):
                        c_result: dict = check_cosmic(path=qualified_paths[1])
                        nc_result : dict = check_noncosmic(path=qualified_paths[0])

                    elif bool(re.match(c_prefix_pattern, qualified_paths[0])) and \
                            bool(re.match(nc_prefix_pattern, qualified_paths[1])):
                        c_result : dict = check_cosmic(path=qualified_paths[0])
                        nc_result : dict = check_noncosmic(path=qualified_paths[1])
                    else:
                        return {"REQ Num": req_num, "path": qualified_paths[0], "match": False,
                                "note": "Incorrect type of cosmic/non-cosmic excel based on requirement"}

                    return {"REQ Num": req_num, "path": qualified_paths[0], "match": c_result['match'] & nc_result['match'],
                        "note": (c_result['note'] + '; ' + nc_result['note']).rstrip('; ')}

                except KeyError:
                    return {"REQ Num": req_num, "path": qualified_paths[0], "match": False, "note": "KeyError, Check column name"}

            else:
                return {"REQ Num": req_num, "path": qualified_paths[0], "match": False,
                        "note": "Incorrect number of cosmic/non-cosmic excel(s) based on requirement"}

        else:
            return {"REQ Num": req_num, "path": qualified_paths[0], "match": False,
                    "note": f"The parameter {qualified_cosmic} is not accepted"}

    def check_all_files(self, check_final_confirmation: bool = True,
                        check_highlight_cfp: bool = True) -> dict[str, list[dict, None]]:
        '''
        Check all related files listed in the result summary.
        Call `check_file` function for each single check.
        The time complexity is omega(n^2) for calling n items in the Excel of result summary.

        :param check_final_confirmation: bool for whether checking final confirmation, default to True
        :param check_highlight_cfp: bool for whether checking highlighting and corresponding cfp, default to True
        :return: A list of results in dict-format. Could be empty list if nothing found.
        '''

        if self.data_frame_specific is None:
            raise CosmicExcelCheckerException("Specific worksheet is not loaded. Use `set_sheet_name` to load it")

        start_time = time.time()
        # use iterrows() to iterate each row in DataFrame
        list_results : list[dict, None] = []
        for index, row in self.data_frame_specific.iterrows():
            list_results.append(self.check_file(
                req_num=row[RS_REQ_NUM],
                check_final_confirmation=check_final_confirmation,
                check_highlight_cfp=check_highlight_cfp
            ))

        cf_results = {
            "results": list_results,
            "time": round(time.time() - start_time, 5)
        }

        return cf_results

