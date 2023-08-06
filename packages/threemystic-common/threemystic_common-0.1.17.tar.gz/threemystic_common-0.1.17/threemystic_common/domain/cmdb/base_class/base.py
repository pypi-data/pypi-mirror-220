from abc import abstractmethod
import asyncio
from threemystic_common.base_class.base_common import base
from openpyxl import Workbook

class cmdb_base(base): 
  """This is a set of library wrappers to help create a cmdb"""

  def __init__(self, *args, **kwargs) -> None:
    super().__init__(*args, **kwargs)

  def init_workbook(self, *args, **kwargs):
    workbook = Workbook()

    while len(workbook.sheetnames) > 0:
      workbook.remove(workbook[workbook.sheetnames[0]])
    
    return workbook

  def _get_default_columns(self, include_source = False, include_id = False, *args, **kwargs):
    prepend = []
    if include_source:
      prepend.append("Source")
    if include_id:
      prepend.append("Id")

    return {
      "no_region": prepend + ["Account ID", "Account Name"],
      "region": prepend + ["Account ID", "Account Name", "Region" ]
    }

  def _get_default_columns(self, include_source = False, include_id = False, *args, **kwargs):
    prepend = []
    if include_source:
      prepend.append("Source")
    if include_id:
      prepend.append("Id")

    return {
      "no_region": prepend + ["Account ID", "Account Name"],
      "region": prepend + ["Account ID", "Account Name", "Region" ]
    }

  # get_cmdb_source
  @abstractmethod
  def get_source(self, *args, **kwargs):
    raise self._main_reference.exception().exception(
      exception_type = "function"
    ).not_implemented(
      logger = self._main_reference.get_common().get_logger(),
      name = "get_source",
      message = f"Not Implemented"
    )

  def get_report_default_columns(self, data_item, *args, **kwargs):
    if self._main_reference.helper_type().general().is_type(data_item["default_columns"], dict):
      return data_item["default_columns"]["default"]
    
    return data_item["default_columns"]
  
  def process_report_data(self, InventoryDataSheet, report_data, *args, **kwargs):
    for sheet in InventoryDataSheet:
      for data in (report_data[sheet] if not self._main_reference.helper_type().general().is_type(report_data[sheet], asyncio.Task) else report_data[sheet].result()):
        InventoryDataSheet[sheet]["data"].append(
          (self.get_report_default_columns(data_item= data)) +
          [data_item for _, data_item in data["data"].items()] +
          self.generate_tag_columns(InventoryDataSheet=data["tags_columns"]["InventoryDataSheet"], tags= data["tags_columns"]["tags"])
        )
  
  # generate_cmdb_data
  def generate_data(self, inventory_data, InventoryDataSheet, report_data, *args, **kwargs):
    cmdb_data = {
      sheet:[] for sheet in InventoryDataSheet
    }    

    if not inventory_data.config.cmdb.enabled:
      return cmdb_data
    
    for sheet in InventoryDataSheet:
      if report_data.get(sheet) is None:
        continue

      cmdb_data[sheet]+=report_data[sheet] if not self._main_reference.helper_type().general().is_type(report_data[sheet], asyncio.Task) else report_data[sheet].result()     
    
    return cmdb_data
  
  # compare_cmdb_account_ids
  def compare_account_ids(self, account1, account2, *args, **kwargs):
    raise self._main_reference.exception().exception(
      exception_type = "function"
    ).not_implemented(
      logger = self._main_reference.get_common().get_logger(),
      name = "compare_account_ids",
      message = f"Not Implemented"
    )
  
  # compare_cmdb_string_possible_numeric
  def compare_string_possible_numeric(cls, str1, str2, *args, **kwargs):
    str1 = str(str1).strip(" '")
    str2 = str(str2).strip(" '")
    if not str1.lower().startswith("=text("):
      str1 = f'=Text("{str1}", "?")'
    if not str2.lower().startswith("=text("):
      str2 = f'=Text("{str2}", "?")'

    return str1.lower() == str2.lower()

  @abstractmethod
  def generate_resource_tags_csv(self, *args, **kwargs): 
    raise self._main_reference.exception().exception(
        exception_type = "function"
      ).not_implemented(
        logger = self._main_reference.get_common().get_logger(),
        name = "generate_resource_tags_csv",
        message = f"Not Implemented"
      )
  
  # generate_tag_columns_cmdb
  # generate_tag_columns
  @abstractmethod
  def generate_tag_columns(self, *args, **kwargs):
    # This will combine the 2 methods into one. I just would need to pass a variable to indicate if it should contain the source column
    raise self._main_reference.exception().exception(
      exception_type = "function"
    ).not_implemented(
      logger = self._main_reference.get_common().get_logger(),
      name = "generate_tag_columns",
      message = f"Not Implemented"
    )
  
  # generate_tag_columns_cmdb_basic
  # generate_tag_columns_basic
  @abstractmethod
  def generate_tag_columns_basic(self, *args, **kwargs):
    # This will combine the 2 methods into one. I just would need to pass a variable to indicate if it should contain the source column
    raise self._main_reference.exception().exception(
      exception_type = "function"
    ).not_implemented(
      logger = self._main_reference.get_common().get_logger(),
      name = "generate_tag_columns_basic",
      message = f"Not Implemented"
    )
  
  def get_tags_as_dict(self, *args, **kwargs):
    # the tags passed into the generate_tag_columns should be a standardized dictionary of {Key:Value}
    raise self._main_reference.exception().exception(
      exception_type = "function"
    ).not_implemented(
      name = "get_tags_as_dict",
      message = f"get_tags_as_dict should be part of the cloud provider something like get resource_tags_as_dict(self, resource_tags"
    )
  
  def get_custom_require_tags(self, *args, **kwargs):
    if hasattr(self, "_custom_require_tags"):
      return self._custom_require_tags
    
    self._custom_require_tags = {}
    return self.get_custom_require_tags(*args, **kwargs)

  def add_custom_require_tags(self, custom_require_tags, merge_if_existing = True, *args, **kwargs):
    
    if not self._main_reference.helper_type().general().is_type(custom_require_tags, dict):
      raise self._main_reference.exception().exception(
        exception_type = "argument"
      ).type_error(
        logger = self._main_reference.get_common().get_logger(),
        name = "custom_require_tags",
        message = f"custom_require_tags is of type {type(custom_require_tags)}"
      )

    if merge_if_existing:
      self._custom_require_tags = self._main_reference.helper_type().dictionary().merge_dictionary(
        [
          {},
          self.get_custom_require_tags(),
          custom_require_tags
        ]
      )
      return self.get_custom_require_tags()
    
    for custom_tag_key, custom_tag_value in custom_require_tags.items():
      self._custom_require_tags[custom_tag_key] = custom_tag_value
    
    return self.get_custom_require_tags()
    


  
  def remove_custom_require_tags(self, custom_require_tag_key, raise_exception = False, *args, **kwargs):    
    if custom_require_tag_key not in self.get_custom_require_tags():
      if raise_exception:
        raise self._main_reference.exception().exception(
          exception_type = "generic"
        ).key_error(
          logger = self._main_reference.get_common().get_logger(),
          name = "custom_require_tags",
          message = f"key not found: {custom_require_tag_key}"
        )
      return None
    
    return self._custom_require_tags.pop(custom_require_tag_key)

  def required_tag_names(self, *args, **kwargs):
    return self._main_reference.helper_type().dictionary().merge_dictionary(
      [
        {}, 
        {} if self._main_reference.get_configuration().get("cmdb") is None else self._main_reference.get_configuration().get("cmdb").get("require_tags"),
        self.get_custom_require_tags()
      ]
    )
  
  # init_report_workbook_cmdb_header_row
  def init_report_workbook_header_row(self, datasheet, inventory_data_column_info, *args, **kwargs):
    raise self._main_reference.exception().exception(
      exception_type = "function"
    ).not_implemented(
      logger = self._main_reference.get_common().get_logger(),
      name = "compare_account_ids",
      message = f"Not Implemented"
    )

  # init_report_workbook_cmdb
  def init_report_workbook(self, InventoryDataSheet, inventory_data_column_info, lambda_create, include_deleted_column = False, *args, **kwargs):
    raise self._main_reference.exception().exception(
      exception_type = "function"
    ).not_implemented(
      logger = self._main_reference.get_common().get_logger(),
      name = "compare_account_ids",
      message = f"Not Implemented"
    )
  
  def get_workbook_column_header_display_dict(self, column_display, default, existing_header = None, *args, **kwargs):
    if self._main_reference.helper_type().general().is_type(column_display, str):
      return column_display

    if existing_header is not None and "search" in column_display and callable(column_display.get("search")):
      return existing_header if column_display.get("search")(existing_header) else default
    
    if "display" in column_display:
      return column_display.get("display")
    
    raise Exception(f"unknown column_display_dict - {self._main_reference.helper_json().dumps(data= column_display)}")
  
  def get_workbook_column_header_display(self, inventory_data_column_info_column, display_key = "default", existing_header = None, *args, **kwargs):
    if self._main_reference.helper_type().general().is_type(inventory_data_column_info_column["display"], str):
      return inventory_data_column_info_column["display"]

    if self._main_reference.helper_type().general().is_type(inventory_data_column_info_column["display"], dict):      
      if display_key in inventory_data_column_info_column["display"]:      
        return self.get_workbook_column_header_display_dict(column_display = inventory_data_column_info_column["display"].get(display_key), default = inventory_data_column_info_column["display"]["default"], existing_header = existing_header)
        
      return inventory_data_column_info_column["display"]["default"]
    
    raise Exception(f"unknown inventory_data_column_info_column - {self._main_reference.helper_json().dumps(inventory_data_column_info_column)}")
  
  def pre_init_report_workbook(self, *args, **kwargs):
    raise self._main_reference.exception().exception(
      exception_type = "function"
    ).not_implemented(
      logger = self._main_reference.get_common().get_logger(),
      name = "compare_account_ids",
      message = f"Not Implemented"
    )
  
  def init_report_workbook(self, InventoryDataSheet, inventory_data_column_info, lambda_create, *args, **kwargs):
    raise self._main_reference.exception().exception(
      exception_type = "function"
    ).not_implemented(
      logger = self._main_reference.get_common().get_logger(),
      name = "compare_account_ids",
      message = f"Not Implemented"
    )
  
  def get_tag_report_columns(self, InventoryDataSheet, *args, **kwargs):
    raise self._main_reference.exception().exception(
      exception_type = "function"
    ).not_implemented(
      logger = self._main_reference.get_common().get_logger(),
      name = "compare_account_ids",
      message = f"Not Implemented"
    )

  def get_default_report_columns(self, InventoryDataSheet, *args, **kwargs):
    raise self._main_reference.exception().exception(
      exception_type = "function"
    ).not_implemented(
      logger = self._main_reference.get_common().get_logger(),
      name = "compare_account_ids",
      message = f"Not Implemented"
    )

  def get_default_report_columns_cmdb(self, InventoryDataSheet, *args, **kwargs):
    raise self._main_reference.exception().exception(
      exception_type = "function"
    ).not_implemented(
      logger = self._main_reference.get_common().get_logger(),
      name = "compare_account_ids",
      message = f"Not Implemented"
    )

  # get_report_default_row_cmdb
  @abstractmethod
  def get_report_default_row(self, *args, **kwargs):
    raise self._main_reference.exception().exception(
        exception_type = "function"
      ).not_implemented(
        logger = self._main_reference.get_common().get_logger(),
        name = "generate_resource_tags_csv",
        message = f"Not Implemented"
      )
  
  @abstractmethod
  def get_account_environment(self, *args, **kwargs):
    raise self._main_reference.exception().exception(
        exception_type = "function"
      ).not_implemented(
        logger = self._main_reference.get_common().get_logger(),
        name = "generate_resource_tags_csv",
        message = f"Not Implemented"
      )

  # save_report
  def save_report(self, report_dir, report_name, workbook, show_output = True):
    date_now = self._main_reference.helper_type().datetime().get(time_zone = "local")

    report_name = report_name.format(self._main_reference.helper_type().datetime().datetime_as_string(dt_format = "%Y%M%d%H%M", dt = date_now))
    # report_directory_path_info = Path("{}/".format(report_dir)).resolve()    
    report_directory_path_info = self._main_reference.helper_path().get(path = f"{report_dir}/")
    report_path = report_directory_path_info.joinpath(f"{report_name}")
    
    if not report_directory_path_info.exists():
      report_directory_path_info.mkdir(parents=True, exist_ok=True)
    
    if show_output:
      print(f"Report will be saved @ {report_path.absolute()}")
    
    workbook.save(report_path.absolute())
  
  def generate_data_item(self, item, data_key, InventoryDataSheet, inventory_data_column_info, tags = None, raw_item = None):
    if tags is None:
      tags = item.get("Tags") if item.get("Tags") is not None else item.get("tags")
    
    
    if item.get("Tags") is None:
      item["Tags"] = tags

    extra_account = item.get("extra_account")
    extra_region = item.get("extra_region")
    extra_resourcegroups = item.get("extra_resourcegroups")
    if raw_item is not None:
      item["extra_raw_item"] = raw_item
      extra_account = raw_item.get("extra_account")
      extra_region = raw_item.get("extra_region")
      extra_resourcegroups = raw_item.get("extra_resourcegroups")
    
    return {
      "raw_item": item,
      "default_columns": self.get_report_default_row(account=extra_account, resource= raw_item if raw_item is not None else item,  region=extra_region, InventoryDataSheet=InventoryDataSheet[data_key], include_cmdb= include_cmdb, resource_groups = extra_resourcegroups),
      "tags_columns":{
        "InventoryDataSheet": InventoryDataSheet[data_key],
        "tags": tags
      },
      "data":{
        column:inventory_data_column_info[data_key][column]["handler"](item) for column in inventory_data_column_info[data_key]
      }
    }