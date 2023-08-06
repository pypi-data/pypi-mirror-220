# bot-lib-public version 2.0.0
# now have Excellib, MQueues, ImageSensor and Threading capabilities 

import platform
import sys
import os
import zipfile
import pickle
import csv
import time
import random
from queue import Queue
from threading import Thread
import concurrent.futures

def install_dependencies():
    requirements = ['selenium', 'requests']
    for each in requirements:
        os.system(f'pip install {each}')

def exportCSV(filename:str, headers:list, data:list):
    data.insert(0, headers)
    try:
        with open(filename, mode='w', newline='') as file:
            writer = csv.writer(file)
            writer.writerows(data)
    except Exception as e:
        print(f"Failed to write CSV due to exception {e}")

def detect_platform():
    current_system = platform.system()
    if current_system == "Windows":
        return "Windows " + str(platform.architecture()[0])
    elif current_system == "Darwin":
        return "Darwin"
    elif current_system == "Linux":
        return "Linux"


def direct_download(URL, file_path_with_extension, show_output=False, user_agent=None):
    USER_AGENT = '''Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36'''
    import requests
    if user_agent is None:
        headers = {'user-agent': USER_AGENT}
    else:
        headers = {'user-agent': user_agent}
    try:
        r = requests.get(URL, headers=headers, stream=True)
        with open(f'{file_path_with_extension}', "wb") as downloading_file:
            total_size = (int(r.headers.get('Content-Length')) / 1000000)
            progress = 0
            for chunk in r.iter_content(chunk_size=1024):
                if chunk:
                    downloading_file.write(chunk)
                    if show_output is True:
                        sys.stdout.write("\r" + str(
                            (progress / 1000).__round__(2)) + f" Mbs Downloaded out of {total_size.__round__(2)}")
                    progress += 1
        if show_output is True:
            print("\n" + f"Successfully downloaded")
        return file_path_with_extension
    except Exception:
        if show_output is True:
            print("\n" + "Failed downloading driver")
        return None

def dict_reverse(dict_input):
    reversed_char_set = {}
    index = 0
    keys_of_dict = list(dict_input.keys())
    values_of_dict = list(dict_input.values())
    while index < len(keys_of_dict):
        reversed_char_set[values_of_dict[index]] = keys_of_dict[index]
        index += 1
    return reversed_char_set


def random_screen():
    """ Returns a random screen size from the global dict of screen sizes """
    GLOBAL_SCREEN_SIZES = {3: "1920,1080", 2: "1366,768", 1: "1280x720"}
    random_number = random.randint(1, 3)
    return GLOBAL_SCREEN_SIZES[random_number]

def threadPool(values, targetFunction, maxWorkers=None):
    if maxWorkers == None:
        with concurrent.futures.ThreadPoolExecutor(max_workers=maxWorkers) as executor:
            results = list(executor.map(targetFunction, values))
    else:
        with concurrent.futures.ThreadPoolExecutor() as executor:
            results = list(executor.map(targetFunction, values))
    return results

def spawn_thread(function, arguments=(), end_on_exit=False):
    target_thread = Thread(target=function, args=arguments)
    target_thread.start()
    if end_on_exit is True:
        target_thread.isDaemon()
    return target_thread

def kill_thread(thread):
    thread.join()
    return None

def subroutine(function, arguments=None):
    routine_function_return = MQueues.create_queue()

    def deploy(*args, **kwargs):
        if arguments:
            return_ = function(arguments)
        else:
            return_ = function()
        MQueues.add_to_queue(routine_function_return, return_)

    if arguments:
        spawn_thread(deploy, arguments=(arguments,))
    else:
        spawn_thread(deploy)
    return routine_function_return

def test_chrome_driver(driver_file_path):
    class driver_test(Bot):
        def __init__(self, driver_path):
            super().__init__(driver_exe=driver_path,
                             force_spawn_browser=False)
            self.get_url("""https://www.google.com""")

    try:
        driver_test(driver_file_path)
        return "WORKING"
    except Exception:
        return "FAILED"

def check_driver_presence(driver_path):
    """ Checks Chrome Driver Presence and tests it """
    print("Driver check running")
    result = os.path.isfile(driver_path)
    if result is True:
        check = test_chrome_driver(driver_path)
        if check == "WORKING":
            return "OK"
        if check == "FAILED":
            print("Internet Connection Check Failed")
            return "NOT OK"
    elif result is False:
        return "NOT FOUND"


class Bot:
    """ It is wrapper for selenium based bot to offer ease of use and quick development
        This library uses several default libraries to offer ease of development of bots
        and automated tests."""
    
    USER_AGENT = '''user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36'''

    def __init__(self,
                 force_spawn_browser=True,
                 window_size="1366,768",
                 user_data_path=None,
                 incognito_mode=False,
                 proxy=False,
                 sock5_ip=None,
                 socks5_port=None,
                 target_name=__file__,
                 driver_exe="Drivers/driverchrome.exe",
                 browser="Chrome",
                 default_profile=False,
                 debug=False,
                 noGPU=False,
                 caps=False):

        from selenium import webdriver
        from selenium.common import exceptions
        from selenium.webdriver import ActionChains
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys as Keyboard
        from selenium.webdriver.support import expected_conditions as ec
        from selenium.webdriver.support.ui import WebDriverWait

        self.ec = ec
        self.Exceptions = exceptions
        self.By = By
        self.ActionChains = ActionChains
        self.Keyboard = Keyboard
        self.WebDriverWait = WebDriverWait
        
        """ Spawn a webdriver session """
        # ----------------------------------------> Google Chrome
        if driver_exe is not None:
            if browser.lower() == "chrome":

                from selenium.webdriver.chrome.service import Service as ChromeService

                options = self.bot_options(browser=browser)
                if proxy is True:
                    options.add_argument("--proxy-server=socks5://" + str(sock5_ip) + ":" + str(socks5_port))
                options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
                options.add_experimental_option('useAutomationExtension', False)
                options.add_argument('--disable-blink-features=AutomationControlled')
                options.add_argument("--log-level=3")
                if noGPU:
                    options.add_argument("--disable-gpu")
                if debug:
                    options.add_argument("--remote-debugging-port=9222")
                if force_spawn_browser is False:
                    options.add_argument('--headless')
                if incognito_mode is True:
                    options.add_argument('--incognito')
                options.add_argument(f"--window-size={window_size}")
                if default_profile is True:
                    options.add_argument(f"user-data-dir={self.get_user_data_path_chrome()}")
                    user_data_path = None  # So user data path cannot be used
                if user_data_path is None:
                    pass
                else:
                    options.add_argument(f"user-data-dir={user_data_path}")
                options.add_argument(self.USER_AGENT)
                if caps:
                    from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
                    # Launch with or without metrics enabled
                    caps = DesiredCapabilities.CHROME.copy()
                    caps['goog:loggingPrefs'] = {'performance': 'ALL'}
                    self.driver = webdriver.Chrome(service=ChromeService(executable_path=str(driver_exe)), options=options,
                                                desired_capabilities=caps)
                else:
                    self.driver = webdriver.Chrome(service=ChromeService(executable_path=str(driver_exe)), options=options)
            # --------------------------------------> Microsoft Edge
            elif browser.lower() == "edge":
                from selenium.webdriver.edge.service import Service as EdgeService
                options = self.bot_options(browser=browser)
                if proxy is True:
                    options.add_argument("--proxy-server=socks5://" + str(sock5_ip) + ":" + str(socks5_port))
                options.add_experimental_option("useAutomationExtension", False)
                options.add_experimental_option("excludeSwitches", ["enable-automation", 'enable-logging'])
                if force_spawn_browser is False:
                    options.add_argument('--headless')
                if incognito_mode is True:
                    options.add_argument('--incognito')
                options.add_argument(f"--window-size={window_size}")
                options.add_argument("--log-level=3")
                if default_profile is True:
                    options.add_argument(f"user-data-dir={self.get_user_data_path_edge()}")
                    user_data_path = None  # So user data path cannot be used
                if user_data_path is None:
                    pass
                else:
                    options.add_argument(f"user-data-dir={user_data_path}")
                options.add_argument(self.USER_AGENT)
                self.driver = webdriver.Edge(options=options, service=EdgeService(executable_path=str(driver_exe)))
            # --------------------------------------> FireFox Settings
            elif browser.lower() == "firefox":

                from selenium.webdriver import FirefoxProfile
                from selenium.webdriver.firefox.service import Service as FirefoxService

                options = self.bot_options(browser=browser)
                if proxy is True:
                    options.add_argument("--proxy-server=socks5://" + str(sock5_ip) + ":" + str(socks5_port))
                options.set_preference("dom.webdriver.enabled", False)
                options.set_preference('useAutomationExtension', False)
                options.set_preference("webdriver.firefox.marionette", "false")
                if force_spawn_browser is False:
                    options.add_argument('--headless')
                if incognito_mode is True:
                    options.add_argument('--incognito')
                options.add_argument(f"--window-size={window_size}")
                options.add_argument("--log-level=3")  # So user data path cannot be used
                if user_data_path is None:
                    pass
                else:
                    user_data_path = FirefoxProfile(self.get_user_data_path_firefox())
                # Should use default profile if both are turned on
                if default_profile is True:
                    user_data_path = FirefoxProfile(self.get_user_data_path_firefox())
                options.add_argument(self.USER_AGENT)
                self.driver = webdriver.Firefox(firefox_profile=user_data_path,
                                                service=FirefoxService(executable_path=str(driver_exe)),
                                                options=options)
            else:
                raise "No supported driver type selected (Supported are 'Chrome', 'ChromiumEdge', 'Firefox')"
        else:
            raise "Please Specify chromedriver executable to use"

    @staticmethod
    def bot_options(browser):
        if browser.lower() == "chrome":
            from selenium.webdriver.chrome.options import Options as ChromeOptions
            options = ChromeOptions()
        elif browser.lower() == "edge":
            from selenium.webdriver.edge.options import Options as EdgeOptions
            options = EdgeOptions()
        elif browser.lower() == "firefox":
            from selenium.webdriver.firefox.options import Options as FirefoxOptions
            options = FirefoxOptions()
        else:
            raise "Unsupported Browser Options (Supported are 'Chrome', 'Chromium-Edge', 'Firefox')"
        return options

    def if_exists(self, path):
        """ It checks whether an element exists or not on a web page """
        try:
            self.element_by_xpath(path)
            return True
        except self.Exceptions.NoSuchElementException:
            return False

    def contains(self, type, text):
        """ Use this to quickly generate xpath for any web element which contains a
            certain text """
        if self.str_validator(type) is not None:
            if self.str_validator(text) is not None:
                dynamic_xpath = f"""//{type}[contains(text(),""" + text + """)]"""
                return dynamic_xpath

    def input_attribute(self, attribute, value):
        if self.str_validator(attribute) is not None:
            static_xpath = f"""//input[@{attribute}= """ + value + """]"""
            return static_xpath
        else:
            print("Wrong attribute given in self.input_attribute")

    def input_contains(self, value):
        if self.str_validator(value) is not None:
            dynamic_xpath = f"""//input[contains(text(),""" + value + """)]"""
            return dynamic_xpath
        else:
            print("Wrong text contain value given in self.input_contains")

    def at_element(self, type, attribute, name):
        if self.str_validator(type) is not None:
            if self.str_validator(attribute) is not None:
                if self.str_validator(name) is not None:
                    static_xpath = f"""//{type}[@{attribute}=""" + name + """]"""
                    return static_xpath
                else:
                    print("Incomplete Parameters in self.at_element")
            else:
                print("Wrong attribute given in self.at_element")
        else:
            print("Wrong type given in self.at_element")

    def element_by_xpath(self, XPATH_str, minimum_delay=0, max_delay=0):
        """ Find Element on a page using XPATH only """
        self.range_delay(int(minimum_delay), int(max_delay))
        return self.driver.find_element(self.By.XPATH, XPATH_str)

    def elements_by_xpath(self, XPATH_str, minimum_delay=0, max_delay=0):
        """ Find Elements on a page using XPATH only """
        self.range_delay(int(minimum_delay), int(max_delay))
        return self.driver.find_elements(self.By.XPATH, XPATH_str)

    def wait_till_visible(self, Type, locator, minimum_delay=0, max_delay=0, wait=0):
        """ This function helps initiate a function that waits for an element to be clickable
            Enter in XPATH, CLASS_NAME, CSS_SELECTOR as strings don't use By.XPATH here """
        element = None
        if Type == "XPATH":
            element = self.ec.element_to_be_clickable((self.By.XPATH, locator))
        if Type == "CLASS_NAME":
            element = self.ec.element_to_be_clickable((self.By.CLASS_NAME, locator))
        if Type == "CSS_SELECTOR":
            element = self.ec.element_to_be_clickable((self.By.CSS_SELECTOR, locator))
        if Type == "ID":
            element = self.ec.element_to_be_clickable((self.By.ID, locator))
        if Type == "LINK_TEXT":
            element = self.ec.element_to_be_clickable((self.By.LINK_TEXT, locator))
        if Type == "TAG_NAME":
            element = self.ec.element_to_be_clickable((self.By.TAG_NAME, locator))
        if element is not None:
            wait_for_element = self.WebDriverWait(self.driver, wait).until(element)
            try:
                self.range_delay(int(minimum_delay), int(max_delay))
            except Exception:
                pass
            return wait_for_element

    @staticmethod
    def timeout(max_timeout=1000, exceptions=False):
        """ Standard way to create a timeout for bot, MAX_TIMEOUT default value is 100s """
        try:
            time.sleep(max_timeout)
        except KeyboardInterrupt:
            if exceptions is True:
                raise "Timeout forcefully closed"
            else:
                pass
        return None

    def get_url(self, url, minimum_delay=0, max_delay=0, show_output=False):
        """ locate a web page using this function """
        if show_output is True:
            print(f"Driver navigating to {url}")
        try:
            self.range_delay(int(minimum_delay), int(max_delay))
        except Exception:
            pass
        return self.driver.get(str(url))

    def reload_url(self, show_output=False, minimum_delay=0, max_delay=0):
        """ Reload a web page using this function """
        page = self.driver.current_url
        if show_output is True:
            print(f"Driver reloading {page}")
        try:
            self.range_delay(int(minimum_delay), int(max_delay))
        except Exception:
            pass
        return self.driver.get(str(page))

    # New Function
    def max_page_height(self):
        return self.driver.execute_script("return document.getElementsByTagName('HTML')[0].outerHTML.length;")

    # New Function
    def focus_element(self, xpath_of_element):
        """ A lot of elements become un-intractable, so we have to focus them before performing action """
        element = self.element_by_xpath(xpath_of_element)
        self.ActionChains(self.driver).move_to_element(element).perform()

    def scroll_url(self, factor=950):
        """ Does default scroll as compared to the complete scroll on an
            actual mouse scrolled """
        self.driver.execute_script("window.scrollTo(0," + str(factor) + ")")

    def get_screenshot(self, filename):
        screenshot_bytes = self.driver.get_screenshot_as_png()
        with open(filename, "wb") as file:
            file.write(screenshot_bytes)

    def do_action_keyboard(self, Type, locator, KEY, minimum_delay=0, max_delay=0, show_output=False):
        """ This function helps you to perform any action on any element of a webpage """
        try:
            element = None
            if Type == "XPATH":
                element = self.driver.find_element(self.By.XPATH, locator)
            if Type == "CLASS_NAME":
                element = self.driver.find_element(self.By.CLASS_NAME, locator)
            if Type == "CSS_SELECTOR":
                element = self.driver.find_element(self.By.CSS_SELECTOR, locator)
            if Type == "ID":
                element = self.driver.find_element(self.By.ID, locator)
            if Type == "LINK_TEXT":
                element = self.driver.find_element(self.By.LINK_TEXT, locator)
            if Type == "TAG_NAME":
                element = self.driver.find_element(self.By.TAG_NAME, locator)
            if element is not None:
                action = self.ActionChains(self.driver)
                action.send_keys_to_element(element, KEY)
                try:
                    self.range_delay(int(minimum_delay), int(max_delay))
                except Exception:
                    pass
                if show_output is True:
                    print(f"Keyboard Action performed on {locator}")
                return action.perform()
        except Exception:
            if show_output is True:
                print(f"Keyboard Action not performed on {locator}")
            return None

    def while_ctrl(self, Type, locator, KEY, minimum_delay=0, max_delay=0, show_output=False):
        """ This function helps you to perform shortcuts like Ctrl+A etc """
        try:
            element = None
            if Type == "XPATH":
                element = self.driver.find_element(self.By.XPATH, locator)
            if Type == "CLASS_NAME":
                element = self.driver.find_element(self.By.CLASS_NAME, locator)
            if Type == "CSS_SELECTOR":
                element = self.driver.find_element(self.By.CSS_SELECTOR, locator)
            if Type == "ID":
                element = self.driver.find_element(self.By.ID, locator)
            if Type == "LINK_TEXT":
                element = self.driver.find_element(self.By.LINK_TEXT, locator)
            if Type == "TAG_NAME":
                element = self.driver.find_element(self.By.TAG_NAME, locator)
            if element is not None:
                self.ActionChains(self.driver).key_down(self.Keyboard.CONTROL).send_keys(KEY).key_up(
                    self.Keyboard.CONTROL).perform()
                try:
                    self.range_delay(int(minimum_delay), int(max_delay))
                except Exception:
                    pass
                if show_output is True:
                    print(f"Keyboard Action performed on {locator}")
                return True
        except Exception as e:
            if show_output is True:
                print(f"Keyboard Action not performed on {locator}")
            return None

    def do_action_mouse(self, Type, locator, minimum_delay=0, max_delay=0, show_output=False):
        """ This function helps you to perform any action on any element of a webpage """
        try:
            element = None
            if Type == "XPATH":
                element = self.driver.find_element(self.By.XPATH, locator)
            if Type == "CLASS_NAME":
                element = self.driver.find_element(self.By.CLASS_NAME, locator)
            if Type == "CSS_SELECTOR":
                element = self.driver.find_element(self.By.CSS_SELECTOR, locator)
            if Type == "ID":
                element = self.driver.find_element(self.By.ID, locator)
            if Type == "LINK_TEXT":
                element = self.driver.find_element(self.By.LINK_TEXT, locator)
            if Type == "TAG_NAME":
                element = self.driver.find_element(self.By.TAG_NAME, locator)
            if element is not None:
                action = self.ActionChains(self.driver)
                action.click(element)
                try:
                    self.range_delay(int(minimum_delay), int(max_delay))
                except Exception:
                    pass
                if show_output is True:
                    print(f"Click Action performed on {locator}")
                return action.perform()
        except Exception:
            if show_output is True:
                print(f"Click Action not performed on {locator}")
            return None

    def load_cookies(self, PICKLED_COOKIES, show_output=False):
        """ Use this method to load saved cookies from pickle object """
        try:
            cookies = self.object_load(PICKLED_COOKIES)
            for cookie in cookies:
                self.driver.add_cookie(cookie)
            if show_output is True:
                print("Successfully loaded cookies")
            return True
        except Exception:
            if show_output is True:
                print("Failed to load cookies")
            return False

    def save_cookies(self, COOKIE_SAVE_NAME, show_output=False):
        """ Use this method to Save cookies to pickle object """
        try:
            if show_output is True:
                print("Successfully saved cookies")
            self.object_dump(self.driver.get_cookies(), COOKIE_SAVE_NAME)
            return COOKIE_SAVE_NAME
        except Exception:
            if show_output is True:
                print("Failed to save cookies")
            return None

    def range_delay(self, min_range=0, max_range=0, adjust_latency=False, show_output=False):
        """ Produces a ranged delay instead of a fixed delay """
        if adjust_latency is True:
            print("Latency calculator not available")
        if show_output is True:
            print(f"Delay range {min_range}s:{max_range}s")
        time_range = random.randint(min_range, max_range)
        time.sleep(time_range)

    def save_as_pdf(self, filename, min_delay=0, max_delay=0, page_range=None):
        """ Only works with Selenium 4.1 + """
        from selenium.webdriver.common.print_page_options import PrintOptions
        import base64
        self.range_delay(min_delay, max_delay)
        try:
            os.mkdir('PDFs')
        except Exception:
            pass
        if page_range is not None:
            print_obj = PrintOptions()
            print_obj.page_ranges = page_range
            PDF_out = self.driver.print_page(print_options=print_obj)
        else:
            PDF_out = self.driver.print_page()
        with open(f'{filename}', 'wb') as PDF_file:
            PDF_file.write(bytes(base64.b64decode(PDF_out)))

    # Only for Legacy Support Same function as above
    def dump_PDF(self, filename, min_delay=0, max_delay=0, page_range=None, no_alert=False):
        """ Only works with Selenium 4.1 + """
        from selenium.webdriver.common.print_page_options import PrintOptions
        if no_alert is False:
            print(
                "This function is not recommended to be used with bot_lib_v3 use save_as_pdf instead for similar results")
        import base64
        self.range_delay(min_delay, max_delay)
        try:
            os.mkdir('PDFs')
        except Exception:
            pass
        if page_range is not None:
            print_obj = PrintOptions()
            print_obj.page_ranges = page_range
            PDF_out = self.driver.print_page(print_options=print_obj)
        else:
            PDF_out = self.driver.print_page()
        with open(f'{filename}', 'wb') as PDF_file:
            PDF_file.write(bytes(base64.b64decode(PDF_out)))

    @staticmethod
    def absolute_file_path_ob(executing_script_name, target_file, show_output=False):
        # Obsolete (Newer method introduced)
        """ Usage self.absolute_file_path(__file__, 'any_file.mp3') in an inherited class
        or use selenium_bot.absolute_file_path(__file__, 'any_file.mp3') in your script file.
        This method is used to complete path behind files present in the main directory. """
        main_script_path = os.path.dirname(os.path.abspath(executing_script_name))
        if show_output is True:
            print(f"absolute path method being used @ {executing_script_name}")
            print(main_script_path + f"/{target_file}")
        return main_script_path + f"/{target_file}"

    @staticmethod
    def abs_path(target_file, show_output=False):
        if os.path.exists(target_file):
            abs_path_generated = os.path.abspath(target_file).replace('\\', '/')
            if show_output is True:
                print(f"absolute path method being used @ {target_file}")
                print(abs_path_generated)
        else:
            return "Not Found"

    @staticmethod
    def get_user_data_path_chrome():
        value = os.getlogin()
        path = f"""C:/Users/{value}/AppData/Local/Google/Chrome/User Data"""
        return path

    @staticmethod
    def get_user_data_path_edge():
        value = os.getlogin()
        path = f"""C:/Users/{value}/AppData/Local/Microsoft/Edge/User Data"""
        return path

    @staticmethod
    def get_user_data_path_firefox():
        """ Targets the default release profile """
        value = os.getlogin()
        path = f"""C:/Users/{value}/AppData/Roaming/Mozilla/Firefox/Profiles/"""
        for each in os.listdir(path):
            if each.__contains__('default-release'):
                return path + each

    def load_object(self, OBJECT_FILENAME=None, show_output=False):
        """ Load any Pickled object without having to use the context manager """
        try:
            if show_output is True:
                print("Successfully loaded the target object")
            return self.object_load(OBJECT_FILENAME)
        except Exception:
            if show_output is True:
                print("Failed to load the target object")
            return False

    def save_object(self, OBJECT_SAVE_NAME, ANY_VALID_OBJECT=None, show_output=False):
        """ Pickle any object easily without having to do context management """
        try:
            self.object_dump(ANY_VALID_OBJECT, OBJECT_SAVE_NAME)
            if show_output is True:
                print("Successfully saved the target object")
            return OBJECT_SAVE_NAME
        except Exception:
            if show_output is True:
                print("Failed to save the target object")
            return None

    @staticmethod
    def direct_print(any_variable="No Variable", string_message=" No Message "):
        """ Directly print any variable instead of converting them to string in a formatted way """
        b = string_message + ' :: ' + str(any_variable)
        print('\r' + b)
        # return print(string_message + ' :: ' + str(any_variable))

    @classmethod
    def direct_download(self, URL, file_path_with_extension, download_primitive=False, show_output=False, user_agent=None):
        USER_AGENT = '''Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36'''
        import requests
        if user_agent is None:
            headers = {'user-agent': USER_AGENT}
        else:
            headers = {'user-agent': user_agent}
        try:
            with open(f'{file_path_with_extension}', 'wb') as file:
                if download_primitive is True:
                    file.write(requests.get(URL, headers=headers).content)
                    file.close()
                else:
                    r = requests.get(URL, headers=headers, stream=True)
                    with open(f'{file_path_with_extension}', "wb") as downloading_file:
                        total_size = (int(r.headers.get('Content-Length')) / 1000000)
                        progress = 0
                        for chunk in r.iter_content(chunk_size=1024):
                            if chunk:
                                downloading_file.write(chunk)
                                if show_output is True:
                                    sys.stdout.write("\r" + str(
                                        progress / 1000) + f"Downloaded at {file_path_with_extension} in MB out of {total_size}")
                                progress += 1
                if show_output is True:
                    print("\n" + f"Successfully downloaded {file_path_with_extension}")
                return file_path_with_extension
        except Exception:
            if show_output is True:
                print("\n" + "Failed to download the target file")
            return None

    @staticmethod
    def cliprogress(output, callcounter):
        """ Operation Print Output, which means if you want to display an indicator """
        if callcounter == 1:
            sys.stdout.write("\r" + "Progress :: " + output)
        else:
            sys.stdout.write("\r" + "Progress :: " + output)

    @staticmethod
    def give_count(interatable, show_output=False):
        """ Enter any object List, String, Integer, Dict and get a count of total elements """
        feed = None
        type_of_interatable = str(type(interatable))
        if type_of_interatable == "<class 'str'>":
            try:
                feed = int(interatable)
            except Exception:
                feed = 1
            if show_output is True:
                print(feed)
        elif type_of_interatable == "<class 'list'>":
            feed = len(interatable)
            if show_output is True:
                print(feed)
        elif type_of_interatable == "<class 'int'>":
            feed = interatable
            if show_output is True:
                print(feed)
        elif type_of_interatable == "<class 'set'>":
            feed = len(interatable)
            if show_output is True:
                print(feed)
        elif type_of_interatable == "<class 'dict'>":
            feed = len((interatable.keys()))
            if show_output is True:
                print(feed)
        if str(type(feed)) == "<class 'int'>":
            return feed

    @staticmethod
    def object_load(file_name):
        object_loaded = pickle.load(open(file_name, "rb"))
        return object_loaded

    @staticmethod
    def object_dump(any_object, file_name):
        pickle.dump(any_object, open(file_name, "wb"))
        return 'Successful Dump'

    @staticmethod
    def object_purge_load(file_name):
        object_loaded = pickle.load(open(file_name, "rb"))
        os.remove(file_name)
        return object_loaded

class Browsers:
    """ Only for Windows """

    def __init__(self) -> None:
        pass

    @property
    def sample_usage(self):
        """
        Here is a sample usage to implement this in your scripts \n
        import browser_detection \n
        Installed_browsers = Browsers() \n
        print(Installed_browsers.is_chrome_installed) \n

        What is returned?

            {'Registered': True, 'Browser': 'Chrome', 'Version': '103.0.5060.134'}
            or
            {'Registered': False, 'Browser': 'Chrome'}

        ! A word of advise !

        This library only checks for the version of certain browsers from the registry.
        It is very much possible that a working chrome exists not installed properly on
        the machine so, be very careful in using this. If possible only use this to while
        working with browsers installed and registered on a user's machine.

        Thanks \n
        Fahim
        08-August-2022

        """
        var = """ 
        Here is a sample usage to implement this in your scripts 
        import browser_detection 
        Installed_browsers = Browsers()
        print(Installed_browsers.is_chrome_installed)

        What is returned?

            {'Registered': True, 'Browser': 'Chrome', 'Version': '103.0.5060.134'}
            or
            {'Registered': False, 'Browser': 'Chrome'}

        ! A word of advise !

        This library only checks for the version of certain browsers from the registry.
        It is very much possible that a working chrome exists not installed properly on
        the machine so, be very careful in using this. If possible only use this to while
        working with browsers installed and registered on a user's machine.

        Thanks
        Fahim
        08-August-2022

        """
        print(var)
        return None

    @property
    def is_chrome_installed(self):
        """ Windows Registry checking function to Detect Chrome Installation """
        try:
            stream_chrome = os.popen(
                'reg query "HKLM\\SOFTWARE\\Wow6432Node\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\Google Chrome"')
            Google_Chrome = stream_chrome.read()[218:].split("\Installer")[0].replace('\n', '')
            return {"Registered": True, "Browser": "Chrome", "Version": Google_Chrome}
        except Exception:
            return {"Registered": False, "Browser": "Chrome"}

    @property
    def is_edge_installed(self):
        """ Windows Registry checking function to Detect Chromium Edge Installation """
        try:
            stream_edge = os.popen(
                'reg query "HKLM\\SOFTWARE\\Wow6432Node\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\Microsoft Edge"')
            Chromium_Edge = stream_edge.read()[(98 + 73):].replace(" ", "").split("Version")[0].replace('\n', '')
            return {"Registered": True, "Browser": "Edge", "Version": Chromium_Edge}
        except Exception:
            return {"Registered": False, "Browser": "Edge"}

    @property
    def is_opera_installed(self):
        """ Might not work since register number after HKU might change on another computer
            Need to implement a fail proof way to detect Opera """
        try:
            User = os.getlogin()
            Default_path = "C:/Users/" + User + "/AppData/Local/Programs/Opera/"
            opera_version = os.listdir(Default_path)[0]
            return {"Directory Found": True, "Browser": "Opera", "Version": opera_version}
        except Exception:
            return {"Directory Found": False, "Browser": "Opera"}

    @property
    def is_firefox_installed(self):
        try:
            Default_installation_path = r"C:/Program Files/Mozilla Firefox/application.ini"
            with open(Default_installation_path, 'r') as file:
                data = file.read()
                file.close()
            value_start = data.index("Version") + len("version") + 1
            value_end = data.index("BuildID")
            firefox_version = data[value_start:value_end].replace('\n', '')
            return {"Directory Found": True, "Browser": "FireFox", "Version": firefox_version}
        except Exception:
            return {"Directory Found": False, "Browser": "FireFox"}


class FireFox_Driver:
    # Need control flow for when the target browser is not installed or accessible
    def __init__(self, GECKO_VER='31.0'):
        platform_ = platform.system()
        if platform_ == "Windows":
            self.GECKO_VER = GECKO_VER
            try:
                value = os.path.exists("Drivers/driverfox.exe")
                if value is True:
                    pass
                else:
                    self.driver_manager_firefox()
            except Exception:
                pass
        else:
            raise BaseException("Automatic Driver downloads only work with windows systems")

    def __str__(self):
        return "Drivers/driverfox.exe"

    def get_url(self):
        if detect_platform() == 'Windows 64bit':
            return f'https://github.com/mozilla/geckodriver/releases/download/v0.{self.GECKO_VER}/geckodriver-v0.{self.GECKO_VER}-win64.zip'
        elif detect_platform() == 'Windows 32bit':
            return f'https://github.com/mozilla/geckodriver/releases/download/v0.{self.GECKO_VER}/geckodriver-v0.{self.GECKO_VER}-win32.zip'
        elif detect_platform() == 'Linux':
            return f'https://github.com/mozilla/geckodriver/releases/download/v0.{self.GECKO_VER}/geckodriver-v0.{self.GECKO_VER}-linux32.tar.gz'

    def driver_manager_firefox(self):
        try:
            os.mkdir("Drivers")
        except Exception:
            pass
        url = self.get_url()
        output = direct_download(URL=url, file_path_with_extension="Drivers\driverdown.zip", show_output=True)
        if output is not None:
            with zipfile.ZipFile("Drivers\driverdown.zip", "r") as zip_ref:
                zip_ref.extractall("Drivers")
            os.remove("Drivers\driverdown.zip")
            try:
                os.rename('Drivers\geckodriver.exe', "Drivers\driverfox.exe")
            except FileExistsError:
                pass
            return "Drivers\driverfox.exe"
        else:
            try:
                os.remove("Drivers\driverdown.zip")
            except Exception:
                pass
            return None


class Edge_Driver:
    # Need control flow for when the target browser is not installed or accessible
    def __init__(self):
        platform_ = platform.system()
        if platform_ == "Windows":
            try:
                value = os.path.exists("Drivers/driveredge.exe")
                if value is True:
                    pass
                else:
                    self.driver_manager_edge()
            except Exception:
                pass
        else:
            raise BaseException("Automatic Driver downloads only work with windows systems")

    def __str__(self):
        return "Drivers/driveredge.exe"

    def get_url(self):
        Browser_object = Browsers()
        info = Browser_object.is_edge_installed
        if info is not None:
            if detect_platform() == 'Windows 64bit':
                return f'https://msedgedriver.azureedge.net/{info["Version"]}/edgedriver_win64.zip'
            elif detect_platform() == 'Windows 32bit':
                return f'https://msedgedriver.azureedge.net/{info["Version"]}/edgedriver_win32.zip'
            elif detect_platform() == 'Darwin':
                return f'https://msedgedriver.azureedge.net/{info["Version"]}/edgedriver_mac64.zip'
            elif detect_platform() == 'Linux':
                return f'https://msedgedriver.azureedge.net/{info["Version"]}/edgedriver_linux64.zip'

    def driver_manager_edge(self):
        try:
            os.mkdir("Drivers")
        except Exception:
            pass
        url = self.get_url()
        output = direct_download(URL=url, file_path_with_extension="Drivers\driverdown.zip", show_output=True)
        if output is not None:
            with zipfile.ZipFile("Drivers\driverdown.zip", "r") as zip_ref:
                zip_ref.extract('msedgedriver.exe', 'Drivers')
            os.remove("Drivers\driverdown.zip")
            try:
                os.rename('Drivers\msedgedriver.exe', "Drivers\driveredge.exe")
            except FileExistsError:
                pass
            return "Drivers\driveredge.exe"
        else:
            try:
                os.remove("Drivers\driverdown.zip")
            except Exception:
                pass
            return None


class Chrome_Driver:
    # Need control flow for when the target browser is not installed or accessible
    def __init__(self):
        platform_ = platform.system()
        if platform_ == "Windows":
            import requests
            self.requests = requests
            try:
                value = os.path.exists("Drivers/driverchrome.exe")
                if value is True:
                    pass
                else:
                    self.driver_manager()
            except Exception:
                pass
        else:
            raise BaseException("Automatic Driver downloads only work with windows systems")

    def __str__(self):
        return "Drivers/driverchrome.exe"

    def get_chrome_version(self):
        Browser_object = Browsers()
        info = Browser_object.is_chrome_installed
        return info['Version']

    def get_url(self):
        each_val = self.build_url()
        if detect_platform() == 'Windows 64bit':
            return f"https://chromedriver.storage.googleapis.com/{each_val}/chromedriver_win32.zip"
        elif detect_platform() == 'Windows 32bit':
            return f"https://chromedriver.storage.googleapis.com/{each_val}/chromedriver_win32.zip"
        elif detect_platform() == 'Darwin':
            return f"https://chromedriver.storage.googleapis.com/{each_val}/chromedriver_mac64.zip"
        elif detect_platform() == 'Linux':
            return f"https://chromedriver.storage.googleapis.com/{each_val}/chromedriver_linux64.zip"

    def build_url(self):
        version = self.get_chrome_version()
        print(version)
        headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                          'AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/75.0.3770.142 Safari/537.36'}
        XML = self.requests.get('https://chromedriver.storage.googleapis.com/?delimiter=/&prefix=', headers=headers)
        parsed = XML.text.replace('>', ' ').replace('<', ' ').replace('?', ' ').split(' ')
        endpoints = []
        magic_number = version.split('.')[2]
        for each in parsed:
            try:
                parsed.remove('')
            except Exception:
                pass
            if each.__contains__('http://'):
                pass
            elif each.__contains__('icons/'):
                pass
            else:
                if each[1:].__contains__('/'):
                    each_val = each.replace('/', '')
                    endpoints.append(each_val)
                    converted = each_val.split('.')
                    try:
                        if converted[2] == magic_number:
                            return each_val
                    except Exception:
                        pass

    def driver_manager(self):
        try:
            os.mkdir("Drivers")
        except Exception:
            pass
        url = self.get_url()
        output = direct_download(URL=url, file_path_with_extension="Drivers\driverdown.zip", show_output=True)
        if output is not None:
            with zipfile.ZipFile("Drivers\driverdown.zip", "r") as zip_ref:
                zip_ref.extractall("Drivers")
            os.remove("Drivers\driverdown.zip")
            try:
                os.rename("Drivers\chromedriver.exe", 'Drivers\driverchrome.exe')
            except FileExistsError:
                pass
            return 'Drivers\driverchrome.exe'
        else:
            try:
                os.remove("Drivers\driverdown.zip")
            except Exception:
                pass
            return None
        

class MQueues:
    """ MQueues or managed queues is a class in bot_lib_v3.py which offers easy interface
        to create and use queues with less strict ways in your bot projects """

    @staticmethod
    def listen_until(any_queue_object, show_output=False):
        """ Try and listen to a queue until a value shows up then do something """
        while True:
            try:
                get_value = any_queue_object.get_nowait()
                if get_value is not None:
                    return get_value
            except Exception:
                pass

    @staticmethod
    def listen_print(checking_value):
        """ Very useful function to avoid unwanted printing of null values """
        if checking_value is not None:
            print(checking_value)
        else:
            pass

    @staticmethod
    def listen_check(any_queue_object, show_output=False):
        """ Try and listen to a queue for a value if nothing is there pass """
        try:
            get_value = any_queue_object.get_nowait()
            if get_value is not None:
                return get_value
            else:
                return ""
        except Exception:
            pass

    @staticmethod
    def create_queue(capacity_of_queue=0, queue_id="", show_output=False):
        """ It creates a queue with a max capacity of elements """
        if capacity_of_queue <= 0:
            capacity_of_queue = 0
            if show_output is True:
                print(f"Total capacity of this queue {queue_id} is :: unlimited")
        else:
            if show_output is True:
                print(f"Total capacity of this queue {queue_id} is :: {capacity_of_queue}")
        return Queue(maxsize=capacity_of_queue)

    @staticmethod
    def add_to_queue(spawned_queue, target_element=None, show_output=False):
        """ It adds something to the Queue provided in the function """
        if spawned_queue.full() is True:
            if show_output is True:
                print(f"Queue Already Full :: MaxSize {spawned_queue.qsize()}")
            return None
        else:
            if target_element is None:
                return None
            if target_element is not None:
                try:
                    if show_output is True:
                        print(str(target_element) + " ::" + " Adding to the Queue")
                    spawned_queue.put(target_element)
                    return f"Successfully Added {target_element}"
                except Exception:
                    if show_output is True:
                        print(str(target_element) + " ::" + " Failed adding to the Queue")
                    return None

    @staticmethod
    def get_first_from_queue(spawned_queue, force_no_wait=False, show_output=False):
        """ Get something specific from the Queue or sequentially using FIFO method """
        try:
            if force_no_wait is True:
                value_from_queue = spawned_queue.get_nowait()
            else:
                value_from_queue = spawned_queue.get()
            if show_output is True:
                print(f"{value_from_queue} is present in the queue")
            return value_from_queue
        except Exception:
            if show_output is True:
                print(f"Nothing in the queue")
            return None

    @staticmethod
    def get_element_from_queue(spawned_queue, get_element, show_output=False):
        """ Get something specific from the Queue directly regardless of order """
        try:
            if spawned_queue.empty() is True:
                raise Exception
            else:
                if show_output is True:
                    sub_function_output = True
                else:
                    sub_function_output = False
                value_from_queue = MQueues.any_element_from_queue(spawned_queue, get_element, sub_function_output)
                if value_from_queue is None:
                    raise Exception
                else:
                    if show_output is True:
                        print(f"{value_from_queue} is present in the queue")
                    return value_from_queue
        except Exception:
            if show_output is True:
                print(f"{get_element} is not present in the queue")
            return None

    @staticmethod
    def queue_check_empty(returned_queue_object, show_output=False):
        """ It checks it the queue is empty or not """
        if show_output is True:
            print("Is Queue Empty: ", returned_queue_object.empty())
        return returned_queue_object.empty()

    @staticmethod
    def any_element_from_queue(spawned_queue, element, show_output=False):
        """ It extracts an element from a queue and refills the queue with data """
        list_of_elements = []
        while spawned_queue.empty() is not True:
            list_of_elements.append(spawned_queue.get())
        print(list_of_elements)
        return_value = None
        for element_from_list in list_of_elements:
            if element == element_from_list:
                index_of_value = list_of_elements.index(element)
                list_of_elements.pop(index_of_value)
                return_value = element
                if show_output is True:
                    print(f"Extracted element :: {return_value}")
            else:
                pass
        for items in list_of_elements:
            MQueues.add_to_queue(spawned_queue, items)
        return return_value

    @staticmethod
    def remove_element_from_queue(spawned_queue, element, show_output=False):
        """ It removes an element from a queue and refills the queue with the previous data """
        list_of_elements = []
        while spawned_queue.empty() is not True:
            list_of_elements.append(spawned_queue.get_nowait())

        for element_from_list in list_of_elements:
            if element == element_from_list:
                index_of_value = list_of_elements.index(element)
                list_of_elements.pop(index_of_value)
                return_value = element
                if show_output is True:
                    print(f"Removed element :: {return_value}")
            else:
                pass
        for items in list_of_elements:
            MQueues.add_to_queue(spawned_queue, items)
        return None

    @staticmethod
    def queue_check_full(returned_queue_object, show_output=False):
        """ It checks it the queue is Full or not """
        if show_output is True:
            print("Is Queue Full: ", returned_queue_object.full())
        return returned_queue_object.full()


class Excel:
    """ Usage-:
    file = Excel.open("Video Data.xlsx")
    ws = file.active
    ws.cell(row=3, column=7, value="Done")
    file.save("Video Data.xlsx")"""

    @staticmethod
    def open(xlsx_name,
             open_read_only=False,
             sheet_vba=False,
             file_data_only=False,
             enable_links=False):
        import openpyxl
        excel_file_object = openpyxl.open(filename=xlsx_name,
                                          read_only=open_read_only,
                                          keep_vba=sheet_vba,
                                          data_only=file_data_only,
                                          keep_links=enable_links)
        return excel_file_object

    class File:
        @staticmethod
        def new():
            import openpyxl
            " Creates an empty file in the ram"
            try:
                new_file_empty = openpyxl.Workbook()
                return new_file_empty
            except:
                from openpyxl import Workbook
                new_file_empty = Workbook()
                return new_file_empty

        @staticmethod
        def load(xlsx_name, open_read_only=False, sheet_vba=False, file_data_only=False, enable_links=False):
            import openpyxl
            excel_file_object = openpyxl.open(filename=xlsx_name,
                                              read_only=open_read_only,
                                              keep_vba=sheet_vba,
                                              data_only=file_data_only,
                                              keep_links=enable_links)
            return excel_file_object

    @classmethod
    def file_instance(self, name=None):
        import openpyxl
        """ Creates an empty file on the disk and returns the object, if file is already existing
            it will open it instead of giving an error"""
        if isinstance(name, str):
            new_file_object = openpyxl.Workbook()
            if name.split('.')[1] == 'xlsx':
                if os.path.exists(name):  # New addition
                    return self.open(xlsx_name=name)
                else:
                    new_file_object.save(name)
                    return new_file_object
            else:
                raise "Wrong file extension provided must be xlsx"
        elif name is None:
            return openpyxl.Workbook()

    # Legacy Support
    @classmethod
    def create(self, name=None, no_alert=False):
        import openpyxl
        """ Creates an empty file on the disk and returns the object, if file is already existing
            it will open it instead of giving an error"""
        if no_alert is True:
            print(
                "'create' function for Excel module has been deprecated and it is recommended to use 'file_instance' method")
        if isinstance(name, str):
            new_file_object = openpyxl.Workbook()
            if name.split('.')[1] == 'xlsx':
                if os.path.exists(name):  # New addition
                    return self.open(xlsx_name=name)
                else:
                    new_file_object.save(name)
                    return new_file_object
            else:
                raise "Wrong file extension provided must be xlsx"
        elif name is None:
            return openpyxl.Workbook()
        
class ImageSensor:
    def __init__(self, mode='RGBA') -> None:
        from PIL import Image
        from PIL import ImageChops
        import numpy as np
        import math
        self.Image = Image
        self.ImageChops = ImageChops
        self.np = np
        self.math = math

        ''' 
        
        Difference-based Image Sensor which allows users to detect how far one image is from another in terms of similarity from 0.0 perfect copy to onwards crude copy
        Requirements Pillow, Numpy
        
        Sample Usage -:
        sensorObj = ImageSensor()
        image1 = sensorObj.Imageload('IMAGE1')
        image2 = sensorObj.Imageload('IMAGE2')
        match = sensorObj.NPdiff(image1, image2, minFactor=1.0, maxFactor=3.0)
        
        '''
        self.commonImageMode = mode

    def Imageload(self, image, resize=False, resolution=(0, 0)):
        if resize is True:
            return self.Image.open(image).resize(resolution)
        elif resize is False:
            return self.Image.open(image)

    def RMSdiff(self, im1, im2, significance=0.0):
        "Calculate the root-mean-square difference between two images"
        diff = self.ImageChops.difference(im1.convert(self.commonImageMode), im2.convert(self.commonImageMode))
        h = diff.histogram()
        sq = (value * (idx ** 2) for idx, value in enumerate(h))
        sum_of_squares = sum(sq)
        rms = self.math.sqrt(sum_of_squares / float(im1.size[0] * im1.size[1]))
        return rms

    def NPdiff(self, im1, im2, minFactor=1.0, maxFactor=3.0):
        "Calculate the numpy array difference between two images"
        diff = self.ImageChops.difference(im1.convert(self.commonImageMode), im2.convert(self.commonImageMode))
        value = self.np.mean(self.np.array(diff))
        if value == 0.0:
            return "Images Match Perfect", value
        elif value <= minFactor:
            return "Images Match Below factor", value
        elif value <= maxFactor:
            return "Images Match Above factor", value
        elif value > maxFactor:
            return "Images Don't Match", value

if __name__ == '__main__':
    pass