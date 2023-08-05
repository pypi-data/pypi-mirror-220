"""
Created on Mon May 23 20:43:56 2022

@author: Fauston 
"""

import numpy as np 
import xlsxwriter
import nibabel as nib
from openpyxl.utils.cell import get_column_letter
import matplotlib.pyplot as plt

from tfe22540.perso_path import perso_path_string, patient_number_list
from tfe22540.atlas_modif_name import get_corr_atlas_list, get_atlas_list

perso_path, excel_path, subjects_path, patients_path, plot_path, atlas_path = perso_path_string() 

patient_numbers = patient_number_list("string_nb")

def volume_regions_excel(temps, patient_list):
    
    atlas_list = get_atlas_list() 
    atlas_name = get_corr_atlas_list(atlas_list)
    workbook = xlsxwriter.Workbook(excel_path + 'volume_regions.xlsx')
    all_volumes_time = []
    
    for time in temps:
        worksheet = workbook.add_worksheet("Volumes " + time)
        worksheet.write('A1',"Atlas")
        
        all_volumes = np.zeros((len(atlas_name[:,0]),len(patient_list)))
        
        for patient, i, k in zip(patient_list, range(1,len(patient_list)+1), range(len(patient_list))):
            print("Patient"+str(patient))
            print('-------')
            
            colletter = get_column_letter(i+1)
            worksheet.write(colletter+'1',patient)
            
            for name, j, l in zip(atlas_name[:,0:2], range(2,len(atlas_name[:,0])+2), range(len(atlas_name[:,0]))):
                if (i==1):
                    worksheet.write('A'+str(j),name[0])
                
                name_mask = name[1]
                if("Cerebellar/" in name_mask):
                    name_mask = name_mask.replace('Cerebellar/', '')
                if(".nii.gz" in name_mask):
                    name_mask = name_mask.replace('.nii.gz', '')
                if("Cerebelar" in name_mask):
                    name_mask = name_mask.replace('Cerebelar', 'Cerebellar')
                if("Cerebellum/" in name_mask):
                    name_mask = name_mask.replace('Cerebellum/', '')
                if("Harvard/" in name_mask):
                    name_mask = name_mask.replace('Harvard/', '')
                if("Harvard_cortex/" in name_mask):
                    name_mask = name_mask.replace('Harvard_cortex/', '')
                if("Lobes/" in name_mask):
                    name_mask = name_mask.replace('Lobes/', '')
                if("XTRACT/" in name_mask):
                    name_mask = name_mask.replace('XTRACT/', '')
                if("CC/" in name_mask): 
                    name_mask = name_mask.replace('CC/', '')
            
                atlas_a = nib.load(patients_path + "#" + str(patient) + "/Mask/sub" + str(patient) + "_" + name_mask + "_mask_" + time + ".nii.gz")

                atlas = atlas_a.get_fdata()
                
                vol = len(np.where(atlas==1)[0])
        
                worksheet.write(colletter+str(j), vol)
                
                all_volumes[l,k]=vol
                
        all_volumes_time.append(all_volumes)
    workbook.close()
    
    return all_volumes_time

temps = ["E1", "E2"]
all_volumes_time = volume_regions_excel(temps, patient_numbers)
all_volume = np.append(all_volumes_time[0], all_volumes_time[1], axis=1)

signi_atlases = ["Brain-Stem",
                 "Amygdala L",
                 "Amygdala R",
                 "Cerebellum  Crus I R",
                 "Cerebellum  Crus II R",
                 "Cerebellum  I-IV L",
                 "Cerebellum  IX L", 
                 "Cerebellum  VIIIa L",
                 "Cerebellum  VIIIb L",
                 "Cerebellum  VIIIb R",
                 "Cerebellum  VIIb R",
                 "Cerebellum  X L",
                 "Cerebellum  X R", 
                 "Cerebellum Vermis", 
                 "Cortico Ponto Cerebellum  L", 
                 "Cortico Ponto Cerebellum  R",
                 "Fornix L",
                 "Inferior Cerebellar Pedunculus  L", 
                 "Inferior Cerebellar Pedunculus  R",
                 "Hippocampus L",
                 "Hippocampus R",
                 "Middle Cerebellar Peduncle", 
                 "Superior Cerebellar Pedunculus  L",
                 "Superior Cerebellar Pedunculus  R",
                 "Thalamus L",
                 "Thalamus R"]

atlas_list = get_atlas_list()             
atlas_name_raccourci = get_corr_atlas_list(atlas_list)[:,0]
signi_volumes = np.zeros(35*2)
nbsign = 0

for atlas,i in zip(atlas_name_raccourci, range(len(atlas_name_raccourci))):
    if(atlas in signi_atlases):
        nbsign += 1
        signi_volumes = np.append(signi_volumes, all_volume[i,:], axis = 0)
signi_volumes = np.reshape(signi_volumes, ((len(signi_atlases)+1), 35*2))
signi_volumes = np.delete(signi_volumes, (0), axis = 0)        

for j in range(signi_volumes.shape[0]):
    mean_line = np.mean(signi_volumes[j,:][signi_volumes[j,:] != 0])
    signi_volumes[j,:][signi_volumes[j,:] == 0] = mean_line

fig = plt.figure(figsize = (14.5,8))
plt.style.use('seaborn')
plt.boxplot(signi_volumes.T, labels = signi_atlases, vert = False, meanline = True)
plt.xlabel("Volume [#Voxels]")
plt.grid(axis='x', linestyle='--')
plt.grid(axis='y', linestyle='--')
plt.title("Volumes of the conspicuous atlases over the set of patients")
plt.xlim([-5,2500])
# plt.xlim([2500,5000])
fig.tight_layout()
fig.savefig(plot_path + "[Volume] - Volumes of the conspicuous atlases over the set of patients (beginning of the graph).pdf")

fig = plt.figure(figsize = (14.5,8))
plt.style.use('seaborn')
plt.boxplot(signi_volumes.T, labels = signi_atlases, vert = False, meanline = True)
plt.xlabel("Volume [#Voxels]")
plt.grid(axis='x', linestyle='--')
plt.grid(axis='y', linestyle='--')
plt.title("Volumes of the conspicuous atlases over the set of patients")
# plt.xlim([-5,2500])
plt.xlim([2500,5000])
fig.tight_layout()
fig.savefig(plot_path + "[Volume] - Volumes of the conspicuous atlases over the set of patients (end of the graph).pdf")